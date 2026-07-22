import pdfplumber
import pandas as pd
import re
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import load_workbook


COLUNAS_FINAL = [
    "Pastas",
    "Valor",
    "Atraso/Mês",
    "Situação",
    "Garantia",
    "Condomínio",
    "sim/não"
]

fill_header = PatternFill(fill_type="solid", fgColor="1F4E78")
fill_dados = PatternFill(fill_type="solid", fgColor="D9EAF7")
fill_vermelho = PatternFill(fill_type="solid", fgColor="FFC7CE")
fill_total = PatternFill(fill_type="solid", fgColor="D9D9D9")

font_header = Font(size=24, bold=True, color="FFFFFF")
font_dados = Font(size=24, color="000000")
font_vermelha = Font(size=24, bold=True, color="9C0006")
font_total = Font(size=24, bold=True, color="000000")

alignment_padrao = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignment_total_label = Alignment(horizontal="left", vertical="center")
alignment_total_valor = Alignment(horizontal="right", vertical="center")

borda = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def normalizar_texto(txt: str) -> str:
    if not txt:
        return ""
    txt = re.sub(r"[ \t]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()


def agrupar_linhas(words, tolerancia_y=3):
    linhas = []
    for w in sorted(words, key=lambda x: (round(x["top"]), x["x0"])):
        adicionou = False
        for linha in linhas:
            if abs(linha["top"] - w["top"]) <= tolerancia_y:
                linha["words"].append(w)
                adicionou = True
                break
        if not adicionou:
            linhas.append({"top": w["top"], "words": [w]})

    linhas_formatadas = []
    for linha in linhas:
        ws = sorted(linha["words"], key=lambda x: x["x0"])
        texto = " ".join(w["text"] for w in ws)
        texto = normalizar_texto(texto)
        linhas_formatadas.append({
            "top": linha["top"],
            "text": texto
        })

    return sorted(linhas_formatadas, key=lambda x: x["top"])


def extrair_tipo_imovel(bloco_texto, endereco_completo="", cabecalho_pagina=""):
    """
    Identifica se o imóvel é apartamento para marcar Condomínio = SIM.

    IMPORTANTE:
    O cabecalho_pagina pode trazer o endereço do proprietário/administradora.
    Exemplo: "RUA Luiz Gama, 178 APTO 62".
    Esse APTO não pertence ao imóvel do bloco, então NÃO deve ser usado
    para definir condomínio.
    """

    # Pega somente a linha do imóvel, exemplo:
    # Imóvel 001 - RUA Santos Dumont, 3494
    m_linha_imovel = re.search(r"Imóvel\s+\d+\s+-[^\n]*", bloco_texto, re.IGNORECASE)
    linha_imovel = m_linha_imovel.group(0) if m_linha_imovel else ""

    # Usa apenas dados do próprio imóvel, nunca o cabeçalho da página.
    texto_base = f"{linha_imovel} {endereco_completo}".upper()

    if re.search(r"\b(APARTAMENTO|APTO|APT)\b", texto_base):
        return "APARTAMENTO", "SIM"

    if (
        re.search(r"\b(BL|BLOCO)\b", texto_base)
        and re.search(r"\b(APTO|APT|APARTAMENTO)\b", texto_base)
    ):
        return "APARTAMENTO", "SIM"

    if re.search(r"\bCASA\b", texto_base):
        return "CASA", "NÃO"
    elif re.search(r"\bSALA COMERCIAL\b|\bCOMERCIAL\b|\bLOJA\b", texto_base):
        return "COMERCIAL", "NÃO"
    elif re.search(r"\bTERRENO\b|\bLOTE\b", texto_base):
        return "TERRENO", "NÃO"
    elif re.search(r"\bBARRACÃO\b|\bBARRACAO\b", texto_base):
        return "BARRACÃO", "NÃO"
    elif re.search(r"\bSOBRADO\b", texto_base):
        return "SOBRADO", "NÃO"

    return "NÃO IDENTIFICADO", "NÃO"

def limpar_trechos_endereco(texto):
    if not texto:
        return ""

    texto = normalizar_texto(texto)
    texto = re.sub(r"^(ENDEREÇO|ENDERECO|LOCALIZAÇÃO|LOCALIZACAO)\s*[:\-]?\s*", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\bCEP\b[:\-]?\s*\d{5}-?\d{3}", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\s+,", ",", texto)
    texto = re.sub(r",\s*,+", ", ", texto)
    texto = re.sub(r"\s*-\s*", " - ", texto)

    return texto.strip(" ,-")


def linha_parece_endereco(linha):
    linha_up = linha.upper()

    termos_ruins = [
        "SITUAÇÃO", "SITUACAO", "INQUILINO", "PROPRIETÁRIO", "PROPRIETARIO",
        "REPASSE", "TOTAL", "ALUG", "CONDOM", "GARANTIA", "FIADOR",
        "CPF", "CNPJ", "FONE", "TEL", "PÁGINA", "PAGINA", "HISTÓRICO", "HISTORICO"
    ]
    if any(t in linha_up for t in termos_ruins):
        return False

    padrao_logradouro = r"\b(RUA|R\.|AV\.?|AVENIDA|ALAMEDA|TRAVESSA|TV\.?|RODOVIA|ESTRADA|PRAÇA|PRACA)\b"
    return bool(re.search(padrao_logradouro, linha, re.IGNORECASE))


def extrair_endereco_completo(bloco_linhas):
    candidatos = []

    for i, linha in enumerate(bloco_linhas):
        linha_limpa = limpar_trechos_endereco(linha)

        if not linha_limpa:
            continue

        if linha_parece_endereco(linha_limpa):
            partes = [linha_limpa]

            for j in range(i + 1, min(i + 3, len(bloco_linhas))):
                prox = limpar_trechos_endereco(bloco_linhas[j])

                if not prox:
                    continue

                prox_up = prox.upper()

                if re.search(r"\b(SITUAÇÃO|SITUACAO|INQUILINO|REPASSE|TOTAL|ALUG|CONDOM|GARANTIA|FIADOR|HISTÓRICO|HISTORICO)\b", prox_up):
                    break

                if (
                    re.search(r"\b(N[º°o]|NÚMERO|NUMERO|APTO|APT|BLOCO|BL|CASA|SALA|LOJA|FUNDOS|QD|LT|LOTE|QUADRA|BAIRRO|CJ|CONJUNTO|CONDOM[IÍ]NIO)\b", prox, re.IGNORECASE)
                    or re.search(r"^\d+[A-Z\-]?$", prox)
                    or len(prox) <= 60
                ):
                    partes.append(prox)
                else:
                    break

            endereco = ", ".join(dict.fromkeys([p for p in partes if p]))
            endereco = limpar_trechos_endereco(endereco)
            candidatos.append(endereco)

    if candidatos:
        candidatos = sorted(candidatos, key=lambda x: len(x), reverse=True)
        return candidatos[0]

    return ""


def valor_negativo(valor):
    try:
        valor_num = float(str(valor).replace(".", "").replace(",", "."))
        return valor_num < 0
    except Exception:
        return False


def converter_valor_para_float(valor):
    """Converte valores de aluguel para número sem perder campos vazios ou sem informação."""
    if valor is None:
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    valor = str(valor).strip()
    if not valor or valor == "-":
        return valor

    valor = valor.replace("R$", "").replace(" ", "")

    try:
        return float(valor.replace(".", "").replace(",", "."))
    except Exception:
        return valor


def extrair_situacao(bloco_linhas, bloco_texto):
    padrao_primeiro_aluguel = r"\b(1\s*[º°O]?\s*ALUGUEL|PRIMEIRO\s+ALUGUEL)\b"

    for linha in bloco_linhas[:12]:
        linha_up = linha.upper()

        if "SITUAÇÃO" in linha_up or "SITUACAO" in linha_up:
            if re.search(padrao_primeiro_aluguel, linha_up):
                return "1º ALUGUEL"
            if "EM RESCISÃO" in linha_up or "EM RESCISAO" in linha_up:
                return "EM RESCISÃO"
            if "LOCADO" in linha_up:
                return "LOCADO"
            if "BAIXADO" in linha_up:
                return "BAIXADO"
            if "DISPONIVEL" in linha_up or "DISPONÍVEL" in linha_up:
                return "DISPONÍVEL"

    texto_up = bloco_texto.upper()

    if re.search(padrao_primeiro_aluguel, texto_up):
        return "1º ALUGUEL"
    if "EM RESCISÃO" in texto_up or "EM RESCISAO" in texto_up:
        return "EM RESCISÃO"
    if "LOCADO" in texto_up:
        return "LOCADO"
    if "BAIXADO" in texto_up:
        return "BAIXADO"
    if "DISPONIVEL" in texto_up or "DISPONÍVEL" in texto_up:
        return "DISPONÍVEL"

    return ""


def contrato_eh_aj(bloco_texto):
    texto_up = bloco_texto.upper()
    return bool(
        re.search(r"\bAJ\b", texto_up) or
        re.search(r"\bAJ\s+\d{4}\b", texto_up) or
        re.search(r"-\s*AJ\s+\d{4}\b", texto_up)
    )


def extrair_alugueis_do_bloco(bloco_texto):
    linhas = bloco_texto.split("\n")
    alugueis = []

    for linha in linhas:
        linha_original = linha.strip()
        linha_up = linha_original.upper()

        if "ESTORNO" in linha_up:
            continue

        if re.search(r"^\s*ALUG\w*", linha_up):
            valores = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|-", linha_original)
            valor = valores[-1].strip() if valores else ""
            em_atraso = ("*" in linha_original) or ("-" == valor)

            alugueis.append({
                "linha": linha_original,
                "valor": valor,
                "em_atraso": em_atraso
            })

    return alugueis


def linha_eh_lancamento_financeiro(linha_up: str) -> bool:
    chaves = [
        "ALUG",
        "CONDOM",
        "IPTU",
        "COMISS",
        "TX",
        "TAXA",
        "SEGURO",
        "PARCIAL",
        "MULTA",
        "JUROS",
        "ÁGUA",
        "AGUA",
        "LUZ",
        "ENERGIA",
        "GÁS",
        "GAS"
    ]
    return any(ch in linha_up for ch in chaves)


def extrair_debitos_do_bloco(bloco_texto):
    linhas = bloco_texto.split("\n")
    debitos = []

    em_historico = False

    for linha in linhas:
        linha_original = linha.strip()
        linha_up = linha_original.upper()

        if not linha_original:
            continue

        if "HISTÓRICO" in linha_up or "HISTORICO" in linha_up:
            em_historico = True
            continue

        if "TOTAL DO IMÓVEL" in linha_up or "TOTAL DO IMOVEL" in linha_up:
            em_historico = False
            continue

        if not em_historico:
            continue

        if any(x in linha_up for x in ["DÉBITO", "DEBITO", "CRÉDITO", "CREDITO"]):
            continue

        if not linha_eh_lancamento_financeiro(linha_up):
            continue

        valores = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", linha_original)
        if not valores:
            continue

        valor = valores[-1].strip()

        tem_sinal_menos = "-" in linha_original
        eh_debito = tem_sinal_menos or any(
            chave in linha_up for chave in ["CONDOM", "IPTU", "COMISS", "TX", "TAXA", "PARCIAL", "MULTA", "JUROS"]
        )

        if eh_debito:
            debitos.append({
                "linha": linha_original,
                "valor": valor
            })

    return debitos


def extrair_creditos_do_bloco(bloco_texto):
    linhas = bloco_texto.split("\n")
    creditos = []

    em_historico = False

    for linha in linhas:
        linha_up = linha.upper()

        if "HISTÓRICO" in linha_up or "HISTORICO" in linha_up:
            em_historico = True
            continue

        if "TOTAL DO IMÓVEL" in linha_up or "TOTAL DO IMOVEL" in linha_up:
            em_historico = False
            continue

        if not em_historico:
            continue

        valores = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", linha)
        if not valores:
            continue

        if "-" not in linha:
            creditos.append(linha)

    return creditos


def coletar_linhas_pdf_continuo(pdf):
    linhas_todas = []
    codigo_atual = ""

    for pagina_num, pagina in enumerate(pdf.pages, start=1):
        words = pagina.extract_words(use_text_flow=True, keep_blank_chars=False)
        if not words:
            continue

        linhas = agrupar_linhas(words)
        cabecalho_pagina = " ".join(l["text"] for l in linhas[:12])
        cabecalho_pagina = normalizar_texto(cabecalho_pagina)

        codigo_encontrado_na_pagina = ""
        for linha in linhas[:15]:
            m = re.search(r"(\d{5})\s*-\s*", linha["text"])
            if m:
                codigo_encontrado_na_pagina = m.group(1)
                codigo_atual = codigo_encontrado_na_pagina
                break

        for linha in linhas:
            linhas_todas.append({
                "pagina": pagina_num,
                "codigo": codigo_atual,
                "cabecalho_pagina": cabecalho_pagina,
                "texto": linha["text"]
            })

    return linhas_todas


def montar_blocos_continuos(linhas_todas):
    blocos = []
    bloco_atual = None

    for item in linhas_todas:
        texto = item["texto"]

        if re.match(r"^Imóvel\s+\d+\s+-", texto, re.IGNORECASE):
            if bloco_atual:
                blocos.append(bloco_atual)

            bloco_atual = {
                "pagina_inicio": item["pagina"],
                "codigo": item["codigo"],
                "cabecalho_pagina": item["cabecalho_pagina"],
                "linhas": [texto]
            }
            continue

        if re.match(r"^Repasse\b", texto, re.IGNORECASE):
            if bloco_atual:
                bloco_atual["linhas"].append(texto)
                blocos.append(bloco_atual)
                bloco_atual = None
            continue

        if bloco_atual:
            if item["codigo"] and not bloco_atual["codigo"]:
                bloco_atual["codigo"] = item["codigo"]

            bloco_atual["linhas"].append(texto)

    if bloco_atual:
        blocos.append(bloco_atual)

    return blocos


def adicionar_totalizador(ws):
    ultima_linha_dados = ws.max_row
    linha_total = ultima_linha_dados + 1

    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

    ws.cell(row=linha_total, column=1, value="Total")

    if "Valor" in headers:
        col_valor = headers.index("Valor") + 1
        letra_col_valor = ws.cell(row=1, column=col_valor).column_letter
        cell_total_valor = ws.cell(
            row=linha_total,
            column=col_valor,
            value=f"=SUM({letra_col_valor}2:{letra_col_valor}{ultima_linha_dados})"
        )
        cell_total_valor.number_format = 'R$ #,##0.00'

    if "Atraso/Mês" in headers:
        col_qtd = headers.index("Atraso/Mês") + 1
        letra_col_qtd = ws.cell(row=1, column=col_qtd).column_letter
        ws.cell(row=linha_total, column=col_qtd, value=f"=SUM({letra_col_qtd}2:{letra_col_qtd}{ultima_linha_dados})")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=linha_total, column=col)
        cell.fill = fill_total
        cell.font = font_total
        cell.border = borda
        cell.alignment = alignment_total_label if col == 1 else alignment_total_valor

    ws.row_dimensions[linha_total].height = 35


def formatar_aba(ws):
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = alignment_padrao
        cell.border = borda

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = font_dados
            cell.fill = fill_dados
            cell.alignment = alignment_padrao
            cell.border = borda

    headers = [c.value for c in ws[1]]

    if "Valor" in headers:
        col_idx_valor = headers.index("Valor") + 1

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx_valor)

            # Deixa toda célula da coluna Valor previamente configurada como moeda.
            # Assim, inclusive na aba "Sem informação de aluguel", quando o usuário
            # digitar manualmente um valor em uma célula vazia, o Excel já exibirá em R$.
            cell.number_format = 'R$ #,##0.00'

            valor_celula = cell.value
            valor_texto = str(valor_celula).strip() if valor_celula is not None else ""

            # Mantém "-" como indicação de valor não encontrado/em atraso.
            # Qualquer valor válido passa a ser numérico e exibido como moeda.
            if valor_texto and valor_texto != "-" and not valor_texto.startswith("="):
                valor_convertido = converter_valor_para_float(valor_celula)
                if isinstance(valor_convertido, (int, float)):
                    cell.value = valor_convertido
                    valor_celula = valor_convertido

            if valor_texto == "-" or (isinstance(valor_celula, (int, float)) and valor_celula < 0):
                cell.fill = fill_vermelho
                cell.font = font_vermelha

    if "Atraso/Mês" in headers:
        col_idx_qtd = headers.index("Atraso/Mês") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx_qtd)
            if isinstance(cell.value, (int, float)) and cell.value > 1:
                cell.fill = fill_vermelho
                cell.font = font_vermelha

    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 35

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 4, 45)





def configurar_larguras_padrao(ws):
    larguras = {"A": 18, "B": 16, "C": 14, "D": 18, "E": 20, "F": 16, "G": 14}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura


def aplicar_titulo_secao(ws, linha, titulo):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=7)
    cell = ws.cell(row=linha, column=1, value=titulo)
    cell.font = Font(size=18, bold=True, color="000000")
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(top=Side(style="medium"), bottom=Side(style="medium"))
    ws.row_dimensions[linha].height = 35


def aplicar_cabecalho_secao(ws, linha):
    for col, titulo in enumerate(COLUNAS_FINAL, start=1):
        cell = ws.cell(row=linha, column=col, value=titulo)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = alignment_padrao
        cell.border = borda
    ws.row_dimensions[linha].height = 42


def normalizar_garantia(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().upper()
    texto = texto.replace("Á", "A").replace("À", "A").replace("Ã", "A").replace("Â", "A")
    texto = texto.replace("É", "E").replace("Ê", "E")
    texto = texto.replace("Í", "I")
    texto = texto.replace("Ó", "O").replace("Ô", "O").replace("Õ", "O")
    texto = texto.replace("Ú", "U")
    texto = texto.replace("Ç", "C")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def valor_excel_para_float(valor):
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace("R$", "").replace(" ", "")
    if not texto or texto == "-":
        return 0.0
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return 0.0


def obter_linhas_origem_segundo_filtro(workbook):
    """
    Lê todas as abas geradas/ajustadas na ETAPA 1.

    Assim, as alterações manuais feitas nas abas Contratos AJ e
    Contratos em rescisão também são consideradas no segundo filtro.
    Para arquivos antigos, também aceita os nomes antigos das abas.
    Quando a mesma pasta aparece mais de uma vez, a linha das abas
    Contratos AJ/Contratos em rescisão tem prioridade, pois é onde o
    usuário pode ter ajustado as informações após a geração pelo PDF.
    """
    abas_origem = [
        "Sem informação de aluguel",
        "Sem informações de aluguel",
        "Aluguel em atraso",
        "Contratos AJ",
        "Contratos em rescisão",
        "Contratos Rescisão"
    ]
    registros_por_pasta = {}

    for nome_aba in abas_origem:
        if nome_aba not in workbook.sheetnames:
            continue

        ws = workbook[nome_aba]
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if not row or not row[0]:
                continue

            pasta = str(row[0]).strip()
            if not pasta or pasta.upper() == "TOTAL":
                continue

            linha = list(row[:7])

            # Contratos que já estão na aba de rescisão sempre saem com a situação padronizada.
            if nome_aba in ["Contratos em rescisão", "Contratos Rescisão"]:
                linha[3] = "RESCISÃO"

            registro_atual = registros_por_pasta.get(pasta)

            if registro_atual is None:
                registros_por_pasta[pasta] = {"linha": linha, "aba": nome_aba}
                continue

            # Se o registro estiver em uma aba de situação específica,
            # mantém essa versão, pois ela pode conter a garantia editada.
            aba_especifica = nome_aba in ["Contratos AJ", "Contratos em rescisão", "Contratos Rescisão"]
            atual_especifica = registro_atual["aba"] in ["Contratos AJ", "Contratos em rescisão", "Contratos Rescisão"]

            if aba_especifica or not atual_especifica:
                registros_por_pasta[pasta] = {"linha": linha, "aba": nome_aba}

    return [item["linha"] for item in registros_por_pasta.values()]


def situacao_normalizada(valor):
    return normalizar_garantia(valor)


def situacao_eh_primeiro_aluguel(valor):
    """Identifica imóveis em situação 1º aluguel / 1° aluguel / primeiro aluguel."""
    texto = situacao_normalizada(valor)
    texto = texto.replace("º", "O").replace("°", "O")
    return bool(re.search(r"\b(1\s*O?\s*ALUGUEL|PRIMEIRO\s+ALUGUEL)\b", texto))


def chave_ordenacao_pasta(pasta):
    """Ordenação crescente natural: 2 vem antes de 10, preservando o padrão das pastas."""
    partes = re.split(r"(\d+)", str(pasta or "").strip().upper())
    return tuple((0, int(parte)) if parte.isdigit() else (1, parte) for parte in partes)


def ordenar_linhas_resultado(linhas):
    """Ordena as linhas por Pastas em ordem crescente em todas as abas da etapa 2."""
    return sorted(linhas, key=lambda linha: chave_ordenacao_pasta(linha[0]))


def ordenar_dataframe_por_pastas(df):
    """Ordena DataFrames da etapa 1 pela coluna Pastas em ordem crescente natural."""
    if df.empty:
        return df

    registros = df.to_dict("records")
    registros = sorted(registros, key=lambda item: chave_ordenacao_pasta(item.get("Pastas", "")))
    return pd.DataFrame(registros, columns=COLUNAS_FINAL).reset_index(drop=True)


def separar_linhas_segundo_filtro(workbook):
    linhas_origem = obter_linhas_origem_segundo_filtro(workbook)

    contratos_aj = []
    contratos_rescisao = []
    acionar_garantia = []
    prioridade_alto_valor_sem_fianca = []
    prioridade_primeiro_aluguel = []

    for linha in linhas_origem:
        situacao = situacao_normalizada(linha[3])
        garantia = normalizar_garantia(linha[4])
        valor = valor_excel_para_float(linha[1])

        if "ACAO" in situacao:
            contratos_aj.append(linha)

        if "RESCISAO" in situacao:
            contratos_rescisao.append(linha)

        # PARTE SUPERIOR: garantias de crédito LOFT e CREDALUGA.
        if garantia in ["LOFT", "CREDALUGA"]:
            acionar_garantia.append(linha)

        # TERCEIRO CAMPO DA PRIORIDADE: imóveis em situação 1º aluguel.
        if situacao_eh_primeiro_aluguel(linha[3]):
            prioridade_primeiro_aluguel.append(linha)

        # PARTE INFERIOR: FIADOR ou TÍTULO acima de R$ 3.000,00
        # e todos os contratos identificados como SEM FIANÇA.
        eh_fiador_ou_titulo = garantia == "FIADOR" or "TITULO" in garantia
        eh_sem_fianca = "SEM FIANCA" in garantia

        if eh_sem_fianca or (eh_fiador_ou_titulo and valor > 3000):
            prioridade_alto_valor_sem_fianca.append(linha)

    return (
        ordenar_linhas_resultado(contratos_aj),
        ordenar_linhas_resultado(contratos_rescisao),
        ordenar_linhas_resultado(acionar_garantia),
        ordenar_linhas_resultado(prioridade_alto_valor_sem_fianca),
        ordenar_linhas_resultado(prioridade_primeiro_aluguel)
    )


def escrever_aba_tabela(workbook, nome_aba, linhas):
    if nome_aba in workbook.sheetnames:
        del workbook[nome_aba]

    ws = workbook.create_sheet(nome_aba)
    for col, titulo in enumerate(COLUNAS_FINAL, start=1):
        ws.cell(row=1, column=col, value=titulo)

    for row_idx, linha in enumerate(ordenar_linhas_resultado(linhas), start=2):
        for col_idx, valor in enumerate(linha[:7], start=1):
            if col_idx == 2:
                valor = converter_valor_para_float(valor)
            ws.cell(row=row_idx, column=col_idx, value=valor)

    formatar_aba(ws)
    adicionar_totalizador(ws)
    configurar_larguras_padrao(ws)
    ws.freeze_panes = "A2"
    return ws


def escrever_linhas_secao(ws, linha_inicio, linhas):
    if not linhas:
        linhas = [["", "", "", "", "", "", ""]]

    for i, dados in enumerate(ordenar_linhas_resultado(linhas), start=linha_inicio):
        for col, valor in enumerate(dados[:7], start=1):
            if col == 2:
                valor = converter_valor_para_float(valor)
            cell = ws.cell(row=i, column=col, value=valor)
            cell.font = font_dados
            cell.fill = fill_dados
            cell.alignment = alignment_padrao
            cell.border = borda
            if col == 2 and isinstance(valor, (int, float)):
                cell.number_format = 'R$ #,##0.00'
        ws.row_dimensions[i].height = 35

    return linha_inicio + len(linhas)


def criar_aba_prioridade_cobranca(workbook, acionar_garantia, prioridade_alto_valor_sem_fianca, prioridade_primeiro_aluguel):
    nomes_antigos = ["Conferência Final", "Prioridade de Cobrança"]
    for nome in nomes_antigos:
        if nome in workbook.sheetnames:
            del workbook[nome]

    nome_aba = "Prioridade de Cobrança"
    ws = workbook.create_sheet(nome_aba)
    configurar_larguras_padrao(ws)

    linha = 1
    aplicar_titulo_secao(ws, linha, "GARANTIA DE CRÉDITO: LOFT E CREDALUGA")
    linha += 1
    aplicar_cabecalho_secao(ws, linha)
    linha += 1
    linha = escrever_linhas_secao(ws, linha, acionar_garantia)

    linha += 2
    aplicar_titulo_secao(ws, linha, "ACIMA DE R$ 3.000: FIADOR OU TÍTULO | TODOS SEM FIANÇA")
    linha += 1
    aplicar_cabecalho_secao(ws, linha)
    linha += 1
    linha = escrever_linhas_secao(ws, linha, prioridade_alto_valor_sem_fianca)

    linha += 2
    aplicar_titulo_secao(ws, linha, "SITUAÇÃO: 1º ALUGUEL")
    linha += 1
    aplicar_cabecalho_secao(ws, linha)
    linha += 1
    escrever_linhas_secao(ws, linha, prioridade_primeiro_aluguel)

    ws.freeze_panes = "A3"
    workbook.active = workbook.sheetnames.index(nome_aba)

    return {
        "garantia_credito": len(acionar_garantia),
        "prioridade_alto_valor_sem_fianca": len(prioridade_alto_valor_sem_fianca),
        "prioridade_primeiro_aluguel": len(prioridade_primeiro_aluguel)
    }


def aplicar_segundo_filtro(workbook):
    contratos_aj, contratos_rescisao, acionar_garantia, prioridade, primeiro_aluguel = separar_linhas_segundo_filtro(workbook)

    escrever_aba_tabela(workbook, "Contratos AJ", contratos_aj)
    escrever_aba_tabela(workbook, "Contratos em rescisão", contratos_rescisao)
    resultado_prioridade = criar_aba_prioridade_cobranca(workbook, acionar_garantia, prioridade, primeiro_aluguel)

    # Remove nome antigo se existir em arquivos anteriores.
    if "Contratos Rescisão" in workbook.sheetnames:
        del workbook["Contratos Rescisão"]

    return {
        "contratos_aj": len(contratos_aj),
        "contratos_rescisao": len(contratos_rescisao),
        **resultado_prioridade
    }


def atualizar_conferencia_excel(caminho_excel_entrada: str, caminho_excel_saida: str = None) -> dict:
    if not caminho_excel_saida:
        caminho_excel_saida = caminho_excel_entrada

    workbook = load_workbook(caminho_excel_entrada)
    resultado = aplicar_segundo_filtro(workbook)
    workbook.save(caminho_excel_saida)
    return {"arquivo_atualizado": caminho_excel_saida, **resultado}


def processar_pdf(arquivo_pdf: str, caminho_saida: str, progress_callback=None, log_callback=None) -> dict:
    """
    ETAPA 1:
    Lê o PDF e gera diretamente as quatro abas de tratamento:
    - Sem informação de aluguel
    - Aluguel em atraso
    - Contratos AJ, para os contratos identificados como AÇÃO/AJ
    - Contratos em rescisão, para os contratos em situação de rescisão

    A aba Prioridade de Cobrança passa a ter três seções:
    - Garantia de crédito: LOFT e CREDALUGA
    - Acima de R$ 3.000: FIADOR ou TÍTULO | todos SEM FIANÇA
    - Situação: 1º aluguel

    Depois da conferência/edição das garantias em qualquer dessas abas,
    a ETAPA 2 relê as informações modificadas, atualiza as abas de
    contratos e cria a aba Prioridade de Cobrança.
    """
    sem_informacao_aluguel = []
    aluguel_em_atraso = []
    contratos_aj = []
    contratos_rescisao = []

    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    def progresso(valor):
        if progress_callback:
            progress_callback(valor)

    progresso(5)
    log("Abrindo PDF...")

    with pdfplumber.open(arquivo_pdf) as pdf:
        linhas_todas = coletar_linhas_pdf_continuo(pdf)
        blocos = montar_blocos_continuos(linhas_todas)

    total_blocos = len(blocos)

    for i, bloco in enumerate(blocos, start=1):
        bloco_linhas = bloco["linhas"]
        bloco_texto = "\n".join(bloco_linhas)
        codigo = bloco["codigo"]

        m_imovel = re.search(r"Imóvel\s+(\d+)", bloco_linhas[0], re.IGNORECASE)
        num_imovel = m_imovel.group(1).zfill(3) if m_imovel else "000"

        m_inq = re.search(r"Inquilino\s+(\d+)", bloco_texto, re.IGNORECASE)
        num_inq = m_inq.group(1).zfill(2) if m_inq else "00"

        pasta = f"{codigo}.{num_imovel}.{num_inq}" if codigo else f"SEM.{num_imovel}.{num_inq}"

        endereco_completo = extrair_endereco_completo(bloco_linhas)
        situacao = extrair_situacao(bloco_linhas, bloco_texto)
        tem_inquilino = num_inq != "00"

        alugueis = extrair_alugueis_do_bloco(bloco_texto)
        tem_aluguel = len(alugueis) > 0
        alugueis_em_atraso = [a for a in alugueis if a["em_atraso"]]
        qtd_alugueis_atraso = len(alugueis_em_atraso)
        valor_aluguel = alugueis[0]["valor"] if alugueis else ""

        debitos = extrair_debitos_do_bloco(bloco_texto)
        tem_debito = len(debitos) > 0
        creditos = extrair_creditos_do_bloco(bloco_texto)
        tem_credito = len(creditos) > 0

        _, eh_apartamento = extrair_tipo_imovel(
            bloco_texto=bloco_texto,
            endereco_completo=endereco_completo
        )

        eh_aj = contrato_eh_aj(bloco_texto)
        # Quando o PDF traz “1º aluguel”, ele deve entrar na rotina como contrato locado,
        # mas a informação precisa ser preservada na coluna Situação para depois alimentar
        # o 3º bloco da aba Prioridade de Cobrança.
        eh_primeiro_aluguel = situacao_eh_primeiro_aluguel(situacao) or situacao_eh_primeiro_aluguel(bloco_texto)
        situacao_operacional = "LOCADO" if eh_primeiro_aluguel else situacao
        situacao_saida = "1º ALUGUEL" if eh_primeiro_aluguel else ("AÇÃO" if eh_aj else situacao)

        entrou_em_regra_atraso = qtd_alugueis_atraso > 0
        entrou_em_regra_sem_info = (not tem_aluguel) and (not tem_debito)
        entrou_em_regra_debito = not tem_aluguel and tem_debito and not tem_credito

        condominio = "SIM" if eh_apartamento == "SIM" else "NÃO"

        registro_base = {
            "Pastas": pasta,
            "Valor": valor_aluguel if valor_aluguel else "",
            "Atraso/Mês": qtd_alugueis_atraso,
            "Situação": situacao_saida,
            "Garantia": "FIADOR",
            "Condomínio": condominio,
            "sim/não": "SIM"
        }
        registro_sem_info = {
            "Pastas": pasta,
            "Valor": "",
            "Atraso/Mês": 0,
            "Situação": situacao_saida,
            "Garantia": "FIADOR",
            "Condomínio": condominio,
            "sim/não": "SIM"
        }

        if situacao_operacional == "LOCADO" and not tem_inquilino:
            if not (entrou_em_regra_atraso or entrou_em_regra_debito or entrou_em_regra_sem_info):
                progresso(min(90, int((i / max(total_blocos, 1)) * 80) + 5))
                continue

        if situacao_operacional == "BAIXADO" and not tem_inquilino:
            if not (entrou_em_regra_atraso or entrou_em_regra_debito):
                progresso(min(90, int((i / max(total_blocos, 1)) * 80) + 5))
                continue

        def incluir_em(lista_destino, situacao_forcada=None):
            if entrou_em_regra_atraso or entrou_em_regra_debito:
                registro = registro_base.copy()
            elif entrou_em_regra_sem_info:
                registro = registro_sem_info.copy()
            else:
                return

            if situacao_forcada:
                registro["Situação"] = situacao_forcada

            lista_destino.append(registro)

        # Contratos AJ/ação já saem diretamente na aba própria do PDF.
        if eh_aj:
            if situacao_operacional == "EM RESCISÃO":
                incluir_em(contratos_rescisao, "RESCISÃO")
            elif situacao_operacional in ["LOCADO", "DISPONÍVEL", "DISPONIVEL"]:
                incluir_em(contratos_aj)
            elif situacao_operacional == "BAIXADO" and (entrou_em_regra_atraso or entrou_em_regra_debito):
                aluguel_em_atraso.append(registro_base)

        # Contratos sem AJ seguem para as demais abas conforme situação.
        elif situacao_operacional in ["LOCADO", "DISPONÍVEL", "DISPONIVEL"]:
            if entrou_em_regra_sem_info:
                sem_informacao_aluguel.append(registro_sem_info)
            if entrou_em_regra_atraso or entrou_em_regra_debito:
                aluguel_em_atraso.append(registro_base)

        elif situacao_operacional == "EM RESCISÃO":
            incluir_em(contratos_rescisao, "RESCISÃO")

        elif situacao_operacional == "BAIXADO":
            if entrou_em_regra_atraso or entrou_em_regra_debito:
                aluguel_em_atraso.append(registro_base)

        progresso(min(90, int((i / max(total_blocos, 1)) * 80) + 5))
        log(f"Bloco {i}/{total_blocos} processado.")

    log("Montando planilhas...")
    progresso(92)

    df_sem_info = pd.DataFrame(sem_informacao_aluguel, columns=COLUNAS_FINAL)
    df_atraso = pd.DataFrame(aluguel_em_atraso, columns=COLUNAS_FINAL)
    df_aj = pd.DataFrame(contratos_aj, columns=COLUNAS_FINAL)
    df_rescisao = pd.DataFrame(contratos_rescisao, columns=COLUNAS_FINAL)

    if not df_sem_info.empty:
        df_sem_info = ordenar_dataframe_por_pastas(df_sem_info.drop_duplicates(subset=["Pastas"]))
    if not df_atraso.empty:
        df_atraso = ordenar_dataframe_por_pastas(df_atraso.drop_duplicates(subset=["Pastas"]))
    if not df_aj.empty:
        df_aj = ordenar_dataframe_por_pastas(df_aj.drop_duplicates(subset=["Pastas"]))
    if not df_rescisao.empty:
        df_rescisao = ordenar_dataframe_por_pastas(df_rescisao.drop_duplicates(subset=["Pastas"]))

    for df in [df_sem_info, df_atraso, df_aj, df_rescisao]:
        if not df.empty:
            df["Valor"] = df["Valor"].apply(converter_valor_para_float)

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_sem_info.to_excel(writer, sheet_name="Sem informação de aluguel", index=False)
        df_atraso.to_excel(writer, sheet_name="Aluguel em atraso", index=False)
        df_aj.to_excel(writer, sheet_name="Contratos AJ", index=False)
        df_rescisao.to_excel(writer, sheet_name="Contratos em rescisão", index=False)

        workbook = writer.book
        for nome_aba in [
            "Sem informação de aluguel",
            "Aluguel em atraso",
            "Contratos AJ",
            "Contratos em rescisão"
        ]:
            ws = workbook[nome_aba]
            formatar_aba(ws)
            adicionar_totalizador(ws)
            configurar_larguras_padrao(ws)

    progresso(100)
    log("Excel inicial finalizado. Confira/edite as garantias e execute a ETAPA 2.")

    return {
        "sem_informacao_aluguel": len(df_sem_info),
        "aluguel_em_atraso": len(df_atraso),
        "contratos_aj": len(df_aj),
        "contratos_rescisao": len(df_rescisao),
        "arquivo_saida": caminho_saida
    }



# ==========================
# APP COM INTERFACE GRÁFICA
# ==========================
import os
import sys
import platform
import subprocess
import threading
import tkinter as tk
import webbrowser
from tkinter import filedialog, messagebox, simpledialog, ttk
from datetime import datetime


def abrir_arquivo_padrao(caminho):
    """
    Abre um arquivo no aplicativo padrão do sistema.
    Compatível com Windows, macOS e Linux.
    """
    caminho = os.path.abspath(caminho)

    if not os.path.exists(caminho):
        raise FileNotFoundError(caminho)

    sistema = platform.system()

    if sistema == "Windows":
        os.startfile(caminho)
    elif sistema == "Darwin":  # macOS
        subprocess.Popen(["open", caminho])
    else:  # Linux
        subprocess.Popen(["xdg-open", caminho])


def pasta_config_app():
    """
    Retorna uma pasta de configuração compatível com Windows, macOS e Linux.
    """
    sistema = platform.system()

    if sistema == "Windows":
        base = os.getenv("APPDATA") or os.path.expanduser("~")
    elif sistema == "Darwin":
        base = os.path.join(os.path.expanduser("~"), "Library", "Application Support")
    else:
        base = os.getenv("XDG_CONFIG_HOME") or os.path.join(os.path.expanduser("~"), ".config")

    pasta = os.path.join(base, "FiltroPDF_Emanuel")
    os.makedirs(pasta, exist_ok=True)
    return pasta


class AppFiltroPDF:
    def __init__(self, root):
        self.root = root
        self.root.title("Filtro de PDF para Excel | Módulo de Anomalias Integrado")
        self.root.geometry("1220x960")
        self.root.configure(bg="#F5F6F7")

        self.pdf_var = tk.StringVar()
        self.excel_var = tk.StringVar()
        self.excel_entrada_var = tk.StringVar()
        self.excel_saida_var = tk.StringVar()

        self.cor_azul = "#1F4E78"
        self.cor_verde = "#2F9254"
        self.cor_cinza = "#6C757D"
        self.cor_fundo = "#F5F6F7"
        self.cor_caixa = "#FFFFFF"
        self.cor_laranja = "#C47F00"

        self.montar_tela()

    def montar_tela(self):
        titulo = tk.Label(self.root, text="Filtro de PDF para Excel", font=("Arial", 28, "bold"), fg=self.cor_azul, bg=self.cor_fundo)
        titulo.pack(pady=(18, 4))

        subtitulo = tk.Label(
            self.root,
            text="Etapa 1: gere as abas do PDF, incluindo AJ e rescisão. Etapa 2: após editar as garantias, atualize essas abas e gere a Prioridade de Cobrança.",
            font=("Arial", 12), fg="#555555", bg=self.cor_fundo
        )
        subtitulo.pack(pady=(0, 14))

        frame1 = tk.Frame(self.root, bg=self.cor_caixa, highlightbackground="#333333", highlightthickness=1)
        frame1.pack(fill="x", padx=18, pady=(5, 10))

        tk.Label(frame1, text="ETAPA 1 - Gerar Excel a partir do PDF", font=("Arial", 15, "bold"), fg=self.cor_azul, bg=self.cor_caixa).grid(row=0, column=0, columnspan=2, sticky="w", padx=20, pady=(14, 8))
        tk.Label(frame1, text="Arquivo PDF", font=("Arial", 13, "bold"), fg=self.cor_azul, bg=self.cor_caixa).grid(row=1, column=0, sticky="w", padx=20, pady=(5, 5))
        tk.Entry(frame1, textvariable=self.pdf_var, font=("Arial", 13), relief="flat", highlightthickness=1, highlightbackground="#AAAAAA").grid(row=2, column=0, sticky="ew", padx=(20, 10), pady=5, ipady=7)
        tk.Button(frame1, text="Buscar PDF", command=self.buscar_pdf, font=("Arial", 11, "bold"), bg=self.cor_azul, fg="white", padx=10, pady=5).grid(row=2, column=1, sticky="ew", padx=(0, 20), pady=5)

        tk.Label(frame1, text="Salvar Excel em", font=("Arial", 13, "bold"), fg=self.cor_azul, bg=self.cor_caixa).grid(row=3, column=0, sticky="w", padx=20, pady=(10, 5))
        tk.Entry(frame1, textvariable=self.excel_var, font=("Arial", 13), relief="flat", highlightthickness=1, highlightbackground="#AAAAAA").grid(row=4, column=0, sticky="ew", padx=(20, 10), pady=(5, 16), ipady=7)
        tk.Button(frame1, text="Escolher local", command=self.escolher_excel, font=("Arial", 11, "bold"), bg=self.cor_azul, fg="white", padx=10, pady=5).grid(row=4, column=1, sticky="ew", padx=(0, 20), pady=(5, 16))
        frame1.grid_columnconfigure(0, weight=1)

        frame2 = tk.Frame(self.root, bg=self.cor_caixa, highlightbackground="#333333", highlightthickness=1)
        frame2.pack(fill="x", padx=18, pady=(0, 10))

        tk.Label(frame2, text="ETAPA 2 - Atualizar contratos e criar Prioridade de Cobrança", font=("Arial", 15, "bold"), fg=self.cor_laranja, bg=self.cor_caixa).grid(row=0, column=0, columnspan=2, sticky="w", padx=20, pady=(14, 8))
        tk.Label(frame2, text="Arquivo Excel editado", font=("Arial", 13, "bold"), fg=self.cor_azul, bg=self.cor_caixa).grid(row=1, column=0, sticky="w", padx=20, pady=(5, 5))
        tk.Entry(frame2, textvariable=self.excel_entrada_var, font=("Arial", 13), relief="flat", highlightthickness=1, highlightbackground="#AAAAAA").grid(row=2, column=0, sticky="ew", padx=(20, 10), pady=5, ipady=7)
        tk.Button(frame2, text="Buscar Excel", command=self.buscar_excel_editado, font=("Arial", 11, "bold"), bg=self.cor_azul, fg="white", padx=10, pady=5).grid(row=2, column=1, sticky="ew", padx=(0, 20), pady=5)

        tk.Label(frame2, text="Salvar novo Excel em", font=("Arial", 13, "bold"), fg=self.cor_azul, bg=self.cor_caixa).grid(row=3, column=0, sticky="w", padx=20, pady=(10, 5))
        tk.Entry(frame2, textvariable=self.excel_saida_var, font=("Arial", 13), relief="flat", highlightthickness=1, highlightbackground="#AAAAAA").grid(row=4, column=0, sticky="ew", padx=(20, 10), pady=(5, 16), ipady=7)
        tk.Button(frame2, text="Escolher local", command=self.escolher_excel_saida, font=("Arial", 11, "bold"), bg=self.cor_azul, fg="white", padx=10, pady=5).grid(row=4, column=1, sticky="ew", padx=(0, 20), pady=(5, 16))
        frame2.grid_columnconfigure(0, weight=1)

        botoes = tk.Frame(self.root, bg=self.cor_fundo)
        botoes.pack(fill="x", padx=18, pady=(4, 8))

        self.btn_processar = tk.Button(botoes, text="1. Processar PDF", command=self.processar_pdf_thread, font=("Arial", 13, "bold"), bg=self.cor_verde, fg="white", width=22, height=2)
        self.btn_processar.pack(side="left", padx=(0, 12))
        self.btn_conferencia = tk.Button(botoes, text="2. Aplicar segundo filtro", command=self.atualizar_conferencia_thread, font=("Arial", 13, "bold"), bg=self.cor_laranja, fg="white", width=28, height=2)
        self.btn_conferencia.pack(side="left", padx=(0, 12))
        self.btn_abrir = tk.Button(botoes, text="Abrir Excel gerado", command=self.abrir_excel, font=("Arial", 13, "bold"), bg=self.cor_cinza, fg="white", width=22, height=2)
        self.btn_abrir.pack(side="left", padx=(0, 12))

        self.btn_admin = tk.Button(botoes, text="Área ADM", command=lambda: exibir_area_admin(self.root), font=("Arial", 13, "bold"), bg="#343A40", fg="white", width=16, height=2)
        self.btn_admin.pack(side="left")

        modulos = tk.Frame(self.root, bg=self.cor_fundo)
        modulos.pack(fill="x", padx=18, pady=(0, 8))
        self.btn_anomalias = tk.Button(
            modulos,
            text="Anomalias / Reajustes e Seguros",
            command=self.abrir_modulo_anomalias,
            font=("Arial", 13, "bold"),
            bg="#087E8B",
            fg="white",
            width=34,
            height=2
        )
        self.btn_anomalias.pack(side="left")

        tk.Label(
            modulos,
            text="Importar o relatório mensal, copiar contratos e abrir o Zimbra.",
            font=("Arial", 11),
            fg="#555555",
            bg=self.cor_fundo
        ).pack(side="left", padx=14)

        self.status = tk.Label(self.root, text="Aguardando arquivo...", font=("Arial", 13, "bold"), bg=self.cor_fundo, fg=self.cor_azul)
        self.status.pack(pady=(8, 5))

        self.progress_frame = tk.Frame(self.root, bg=self.cor_azul, height=30)
        self.progress_frame.pack(fill="x", padx=18, pady=(5, 8))
        self.progress_label = tk.Label(self.progress_frame, text="0%", font=("Arial", 10, "bold"), bg=self.cor_azul, fg="white", anchor="w")
        self.progress_label.pack(fill="both", padx=12)

        tk.Label(self.root, text="Log de processamento", font=("Arial", 13, "bold"), fg=self.cor_azul, bg=self.cor_fundo).pack(anchor="w", padx=20, pady=(5, 5))
        self.log_text = tk.Text(self.root, font=("Consolas", 10), height=10, bg="white", fg="black")
        self.log_text.pack(fill="both", expand=True, padx=18, pady=(0, 15))

    def log(self, msg):
        self.root.after(0, lambda: self._log(msg))

    def _log(self, msg):
        self.log_text.insert("end", str(msg) + "\n")
        self.log_text.see("end")

    def set_status(self, msg, cor=None):
        self.root.after(0, lambda: self.status.config(text=msg, fg=cor or self.cor_verde))

    def set_progress(self, valor):
        self.root.after(0, lambda: self.progress_label.config(text=f"{valor}%"))

    def travar_botoes(self, travar=True):
        estado = "disabled" if travar else "normal"
        self.root.after(0, lambda: [
            self.btn_processar.config(state=estado),
            self.btn_conferencia.config(state=estado),
            self.btn_abrir.config(state=estado),
            self.btn_admin.config(state=estado),
            self.btn_anomalias.config(state=estado)
        ])

    def sugerir_saida_conferencia(self, caminho_entrada):
        if not caminho_entrada:
            return ""
        pasta = os.path.dirname(caminho_entrada)
        nome_sem_ext = os.path.splitext(os.path.basename(caminho_entrada))[0]
        return os.path.join(pasta, f"{nome_sem_ext}_prioridade_cobranca.xlsx")

    def buscar_pdf(self):
        caminho = filedialog.askopenfilename(title="Selecione o PDF", filetypes=[("Arquivos PDF", "*.pdf")])
        if caminho:
            self.pdf_var.set(caminho)
            if not self.excel_var.get():
                pasta = os.path.dirname(caminho)
                nome = f"repasse_aluguel_tratado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                self.excel_var.set(os.path.join(pasta, nome))

    def escolher_excel(self):
        caminho = filedialog.asksaveasfilename(title="Salvar Excel em", defaultextension=".xlsx", filetypes=[("Arquivo Excel", "*.xlsx")])
        if caminho:
            self.excel_var.set(caminho)

    def buscar_excel_editado(self):
        caminho = filedialog.askopenfilename(title="Selecione o Excel editado", filetypes=[("Arquivo Excel", "*.xlsx")])
        if caminho:
            self.excel_entrada_var.set(caminho)
            if not self.excel_saida_var.get():
                self.excel_saida_var.set(self.sugerir_saida_conferencia(caminho))

    def escolher_excel_saida(self):
        caminho = filedialog.asksaveasfilename(title="Salvar novo Excel em", defaultextension=".xlsx", filetypes=[("Arquivo Excel", "*.xlsx")])
        if caminho:
            self.excel_saida_var.set(caminho)

    def processar_pdf_thread(self):
        threading.Thread(target=self.processar_pdf_acao, daemon=True).start()

    def processar_pdf_acao(self):
        pdf = self.pdf_var.get().strip()
        excel = self.excel_var.get().strip()
        if not pdf or not os.path.exists(pdf):
            messagebox.showerror("Erro", "Selecione um arquivo PDF válido.")
            return
        if not excel:
            messagebox.showerror("Erro", "Escolha onde salvar o Excel.")
            return
        try:
            self.travar_botoes(True)
            self.root.after(0, lambda: self.log_text.delete("1.0", "end"))
            self.set_status("Processando PDF...", self.cor_azul)
            self.set_progress(0)
            resultado = processar_pdf(pdf, excel, progress_callback=self.set_progress, log_callback=self.log)
            self.set_status("Processo concluído com sucesso.", self.cor_verde)
            self.log("Processamento concluído com sucesso.")
            self.log(f"Sem informação de aluguel: {resultado['sem_informacao_aluguel']}")
            self.log(f"Aluguel em atraso: {resultado['aluguel_em_atraso']}")
            self.log(f"Contratos AJ: {resultado['contratos_aj']}")
            self.log(f"Contratos em rescisão: {resultado['contratos_rescisao']}")
            self.log("Agora abra o Excel, altere as garantias nas abas necessárias, salve e feche o arquivo.")
            self.log("Depois use a ETAPA 2 para atualizar as abas modificadas e gerar a Prioridade de Cobrança.")
            self.excel_entrada_var.set(excel)
            self.excel_saida_var.set(self.sugerir_saida_conferencia(excel))
        except Exception as e:
            self.set_status("Erro ao processar.", "red")
            messagebox.showerror("Erro", str(e))
        finally:
            self.travar_botoes(False)

    def atualizar_conferencia_thread(self):
        threading.Thread(target=self.atualizar_conferencia_acao, daemon=True).start()

    def atualizar_conferencia_acao(self):
        excel_entrada = self.excel_entrada_var.get().strip()
        excel_saida = self.excel_saida_var.get().strip()
        if not excel_entrada or not os.path.exists(excel_entrada):
            messagebox.showerror("Erro", "Selecione um arquivo Excel editado válido na ETAPA 2.")
            return
        if not excel_saida:
            messagebox.showerror("Erro", "Escolha onde salvar o novo Excel da Prioridade de Cobrança.")
            return
        try:
            self.travar_botoes(True)
            self.set_status("Aplicando segundo filtro e criando Prioridade de Cobrança...", self.cor_azul)
            self.set_progress(10)
            self.log("Lendo Excel editado...")
            self.log(f"Entrada: {excel_entrada}")
            self.log(f"Saída: {excel_saida}")
            resultado = atualizar_conferencia_excel(excel_entrada, excel_saida)
            self.excel_var.set(excel_saida)
            self.set_progress(100)
            self.set_status("Novo Excel com Prioridade de Cobrança criado com sucesso.", self.cor_verde)
            self.log("Segundo filtro e Prioridade de Cobrança criados com sucesso.")
            self.log(f"Contratos AJ: {resultado['contratos_aj']}")
            self.log(f"Contratos em rescisão: {resultado['contratos_rescisao']}")
            self.log(f"Prioridade - garantia LOFT/CREDALUGA: {resultado['garantia_credito']}")
            self.log(f"Prioridade - acima de R$ 3.000 ou sem fiança: {resultado['prioridade_alto_valor_sem_fianca']}")
            self.log(f"Prioridade - situação 1º aluguel: {resultado['prioridade_primeiro_aluguel']}")
            self.log(f"Arquivo criado: {resultado['arquivo_atualizado']}")
        except PermissionError:
            self.set_status("Feche o Excel antes de aplicar o segundo filtro.", "red")
            messagebox.showerror("Erro", "Feche o arquivo Excel antes de criar a Conferência Final.")
        except Exception as e:
            self.set_status("Erro ao aplicar o segundo filtro.", "red")
            messagebox.showerror("Erro", str(e))
        finally:
            self.travar_botoes(False)

    def abrir_excel(self):
        excel = self.excel_saida_var.get().strip() or self.excel_var.get().strip()
        if excel and os.path.exists(excel):
            try:
                abrir_arquivo_padrao(excel)
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível abrir o Excel automaticamente.\n\nArquivo:\n{excel}\n\nErro:\n{e}")
        else:
            messagebox.showwarning("Aviso", "Nenhum Excel válido encontrado para abrir.")

    def abrir_modulo_anomalias(self):
        JanelaAnomalias(self.root)


# ==========================
# MÓDULO INTEGRADO - ANOMALIAS / REAJUSTES / SEGUROS
# ==========================

MSG_ANOMALIA_ALUGUEL = """Olá, tudo bem?

Segue em anexo o relatório com os contratos de aluguel que não tiveram reajuste aplicado no mês vigente, para conferência e tratativas necessárias.

Contratos:
{contratos}

Permaneço à disposição."""

MSG_ANOMALIA_SEGURO_FIANCA = """Olá, tudo bem?

Segue em anexo o relatório com os contratos identificados com seguro fiança vencido no mês vigente, para conferência e regularização necessária.

Contratos:
{contratos}

Permaneço à disposição."""

MSG_ANOMALIA_EMAIL_FINAL = """Olá, tudo bem?

Informo que o relatório de anomalias referente ao mês vigente foi conferido e que as devidas providências foram tomadas em relação aos registros identificados.

Permaneço à disposição."""


def classificar_anomalia_relatorio(mensagem):
    texto = normalizar_texto(mensagem).upper()
    if "ALUGUEL NÃO REAJUSTADO" in texto or "ALUGUEL NAO REAJUSTADO" in texto:
        return "Aluguel não reajustado"
    if "SEGURO FIANÇA" in texto or "SEGURO FIANCA" in texto:
        return "Seguro fiança vencido"
    if "SEGURO INCÊNDIO" in texto or "SEGURO INCENDIO" in texto:
        return "Seguro incêndio vencido/vencendo"
    if "BLOQUEIO" in texto:
        return "Recibo bloqueado para recálculo"
    if "RECIBO MES ANTERIOR NÃO EXISTE" in texto or "RECIBO MÊS ANTERIOR NÃO EXISTE" in texto:
        return "Recibo mês anterior inexistente"
    return "Outros"


def extrair_anomalias_pdf(caminho_pdf):
    with pdfplumber.open(caminho_pdf) as pdf:
        texto = "\n".join((pagina.extract_text() or "") for pagina in pdf.pages)

    texto = re.sub(r"\s+", " ", texto)
    padrao = re.compile(
        r"(?P<horario>\d{2}:\d{2}:\d{2})\s*:\s*"
        r"(?P<mensagem>.*?)\s*"
        r"\[(?P<contrato>\d{5}\.\d{3}\.\d{2})-"
        r"(?P<competencia>\d{1,2}/\d{4})\s+Dia\s+"
        r"(?P<vencimento>\d{1,2})\s+-\s+-"
        r"(?P<garantia>[^\]]+)\]"
    )

    registros = []
    for achado in padrao.finditer(texto):
        mensagem = achado.group("mensagem").strip()
        if "Calculando Local Pagamento" in mensagem and "Aluguel" in mensagem:
            mensagem = mensagem[mensagem.find("Aluguel"):]
        registros.append({
            "horario": achado.group("horario"),
            "categoria": classificar_anomalia_relatorio(mensagem),
            "mensagem": mensagem,
            "contrato": achado.group("contrato"),
            "competencia": achado.group("competencia"),
            "vencimento": achado.group("vencimento"),
            "garantia": achado.group("garantia").strip()
        })
    return registros


def contratos_anomalia_unicos(registros):
    vistos = set()
    contratos = []
    for item in registros:
        if item["contrato"] not in vistos:
            vistos.add(item["contrato"])
            contratos.append(item["contrato"])
    return contratos


def caminho_historico_anomalias():
    pasta = os.path.join(os.getenv("APPDATA") or os.path.expanduser("~"), "FiltroPDF_Emanuel")
    os.makedirs(pasta, exist_ok=True)
    return os.path.join(pasta, "historico_anomalias.json")


def chave_anomalia_registro(registro):
    return (
        registro.get("categoria", ""),
        registro.get("contrato", ""),
        registro.get("competencia", ""),
        registro.get("vencimento", ""),
        registro.get("garantia", ""),
        registro.get("mensagem", "")
    )


def carregar_historico_anomalias():
    import json
    caminho = caminho_historico_anomalias()
    try:
        if os.path.exists(caminho):
            with open(caminho, "r", encoding="utf-8") as arquivo:
                dados = json.load(arquivo)
                if isinstance(dados, list):
                    return dados
    except Exception:
        pass
    return []


def salvar_historico_anomalias(registros):
    import json
    caminho = caminho_historico_anomalias()
    with open(caminho, "w", encoding="utf-8") as arquivo:
        json.dump(registros, arquivo, ensure_ascii=False, indent=4)


def caminho_finalizados_anomalias():
    pasta = os.path.join(os.getenv("APPDATA") or os.path.expanduser("~"), "FiltroPDF_Emanuel")
    os.makedirs(pasta, exist_ok=True)
    return os.path.join(pasta, "finalizados_anomalias.json")


def carregar_finalizados_anomalias():
    import json
    caminho = caminho_finalizados_anomalias()
    try:
        if os.path.exists(caminho):
            with open(caminho, "r", encoding="utf-8") as arquivo:
                dados = json.load(arquivo)
                if isinstance(dados, list):
                    return dados
    except Exception:
        pass
    return []


def salvar_finalizados_anomalias(registros):
    import json
    caminho = caminho_finalizados_anomalias()
    with open(caminho, "w", encoding="utf-8") as arquivo:
        json.dump(registros, arquivo, ensure_ascii=False, indent=4)


def competencias_anomalias(registros):
    competencias = sorted({r.get("competencia", "") for r in registros if r.get("competencia", "")})
    return ["Todas"] + competencias


class JanelaAnomalias(tk.Toplevel):
    CATEGORIAS_GERAIS = [
        "Todos",
        "Aluguel não reajustado",
        "Seguro fiança vencido",
        "Seguro incêndio vencido/vencendo",
        "Recibo bloqueado para recálculo",
        "Recibo mês anterior inexistente",
        "Outros"
    ]
    CATEGORIAS_RENOVACAO = [
        "Todos",
        "Aluguel não reajustado",
        "Seguro fiança vencido"
    ]

    def __init__(self, root):
        super().__init__(root)
        self.title("Anomalias / Reajustes e Seguros | Histórico nos Finalizados | Zimbra")
        self.geometry("1260x760")
        self.configure(bg="#F5F6F7")
        self.minsize(1080, 650)

        self.cor_azul = "#1F4E78"
        self.cor_verde = "#2F9254"
        self.cor_zimbra = "#087E8B"
        self.caminho_pdf = ""
        self.registros = carregar_historico_anomalias()
        self.registros_finalizados = carregar_finalizados_anomalias()
        self.filtro_var = tk.StringVar(value="Todos")
        self.competencia_var = tk.StringVar(value="Todas")
        mensagem_inicial = (
            f"Histórico carregado: {len(self.registros)} ocorrência(s) importada(s) e "
            f"{len(self.registros_finalizados)} ocorrência(s) finalizada(s)."
            if self.registros or self.registros_finalizados else
            "Adicione um ou mais relatórios de anomalias em PDF para montar o histórico."
        )
        self.status_var = tk.StringVar(value=mensagem_inicial)

        self.montar_tela()
        self.atualizar_opcoes_competencia()
        self.atualizar_cards()
        self.atualizar_tabela()

    def montar_tela(self):
        tk.Label(
            self,
            text="Anomalias / Reajustes e Seguros",
            font=("Arial", 23, "bold"),
            fg=self.cor_azul,
            bg="#F5F6F7"
        ).pack(anchor="w", padx=18, pady=(16, 2))

        tk.Label(
            self,
            text="Importe o PDF mensal, filtre os registros e copie as comunicações necessárias.",
            font=("Arial", 11),
            fg="#555555",
            bg="#F5F6F7"
        ).pack(anchor="w", padx=18, pady=(0, 12))

        linha_pdf = tk.Frame(self, bg="white", highlightbackground="#CCCCCC", highlightthickness=1)
        linha_pdf.pack(fill="x", padx=18, pady=(0, 10))

        tk.Button(
            linha_pdf, text="Adicionar PDF ao histórico", command=self.selecionar_pdf,
            font=("Arial", 11, "bold"), bg=self.cor_azul, fg="white", padx=12, pady=8
        ).pack(side="left", padx=12, pady=12)

        tk.Button(
            linha_pdf, text="Limpar histórico", command=self.limpar_historico,
            font=("Arial", 11, "bold"), bg="#A94442", fg="white", padx=12, pady=8
        ).pack(side="left", padx=(0, 12), pady=12)

        texto_pdf = (
            f"Histórico local: {len(self.registros)} ocorrência(s) importada(s) | "
            f"{len(self.registros_finalizados)} finalizada(s)."
            if self.registros or self.registros_finalizados else
            "Nenhum PDF no histórico."
        )
        self.label_pdf = tk.Label(linha_pdf, text=texto_pdf, bg="white", fg="#555555", font=("Arial", 11))
        self.label_pdf.pack(side="left", padx=8)

        cards = tk.Frame(self, bg="#F5F6F7")
        cards.pack(fill="x", padx=18, pady=(0, 10))
        self.card_total = self.criar_card(cards, "Ocorrências")
        self.card_aluguel = self.criar_card(cards, "Não reajustados")
        self.card_seguro = self.criar_card(cards, "Seguro fiança")
        self.card_contratos = self.criar_card(cards, "Contratos únicos")

        acoes = tk.Frame(self, bg="#F5F6F7")
        acoes.pack(fill="x", padx=18, pady=(0, 10))

        tk.Label(acoes, text="Categoria:", font=("Arial", 11, "bold"), bg="#F5F6F7", fg=self.cor_azul).pack(side="left")
        self.combo_filtro_anomalias = ttk.Combobox(
            acoes,
            textvariable=self.filtro_var,
            values=self.CATEGORIAS_GERAIS,
            state="readonly",
            width=28
        )
        self.combo_filtro_anomalias.pack(side="left", padx=(6, 12))
        self.combo_filtro_anomalias.bind("<<ComboboxSelected>>", lambda _e: self.atualizar_tabela())

        tk.Label(acoes, text="Competência:", font=("Arial", 11, "bold"), bg="#F5F6F7", fg=self.cor_azul).pack(side="left")
        self.combo_competencia_anomalias = ttk.Combobox(
            acoes,
            textvariable=self.competencia_var,
            values=["Todas"],
            state="readonly",
            width=12
        )
        self.combo_competencia_anomalias.pack(side="left", padx=(6, 14))
        self.combo_competencia_anomalias.bind("<<ComboboxSelected>>", lambda _e: self.atualizar_tabela())

        self.botao(acoes, "Copiar contrato selecionado", self.copiar_contrato_selecionado, self.cor_azul)
        self.botao(acoes, "Copiar contratos do filtro", self.copiar_contratos_filtro, self.cor_azul)
        self.botao(acoes, "Msg. não reajustado", self.copiar_mensagem_aluguel, self.cor_verde)
        self.botao(acoes, "Msg. seguro fiança", self.copiar_mensagem_seguro, self.cor_verde)
        self.botao(acoes, "FINALIZAR", self.finalizar_selecionados, "#6B2D73")

        # Abas internas: visão geral, visão de renovação e contratos concluídos.
        self.abas_anomalias = ttk.Notebook(self)
        self.abas_anomalias.pack(fill="both", expand=True, padx=18, pady=(0, 10))

        self.aba_geral = tk.Frame(self.abas_anomalias, bg="#F5F6F7")
        self.aba_renovacao = tk.Frame(self.abas_anomalias, bg="#F5F6F7")
        self.aba_finalizados = tk.Frame(self.abas_anomalias, bg="#F5F6F7")
        self.abas_anomalias.add(self.aba_geral, text=" TODAS AS ANOMALIAS ")
        self.abas_anomalias.add(self.aba_renovacao, text=" RENOVAÇÃO ")
        self.abas_anomalias.add(self.aba_finalizados, text=" CONTRATOS FINALIZADOS ")
        self.abas_anomalias.bind("<<NotebookTabChanged>>", self.alterar_filtro_por_aba)

        colunas = ("categoria", "contrato", "competencia", "dia", "garantia", "mensagem")
        titulos = {
            "categoria": "Categoria", "contrato": "Contrato", "competencia": "Competência",
            "dia": "Dia", "garantia": "Garantia", "mensagem": "Mensagem original"
        }
        larguras = {"categoria": 220, "contrato": 125, "competencia": 95, "dia": 58, "garantia": 145, "mensagem": 500}

        area_tabela = tk.Frame(self.aba_geral, bg="#F5F6F7")
        area_tabela.pack(fill="both", expand=True, pady=(8, 0))
        self.tabela = ttk.Treeview(area_tabela, columns=colunas, show="headings")
        for coluna in colunas:
            self.tabela.heading(coluna, text=titulos[coluna])
            self.tabela.column(coluna, width=larguras[coluna], anchor="center" if coluna in ["contrato", "competencia", "dia"] else "w")
        scroll_y = ttk.Scrollbar(area_tabela, orient="vertical", command=self.tabela.yview)
        self.tabela.configure(yscrollcommand=scroll_y.set)
        self.tabela.pack(side="left", fill="both", expand=True)
        scroll_y.pack(side="right", fill="y")

        cabecalho_renovacao = tk.Frame(self.aba_renovacao, bg="white", highlightbackground="#D7DCE1", highlightthickness=1)
        cabecalho_renovacao.pack(fill="x", pady=(8, 8))
        tk.Label(
            cabecalho_renovacao,
            text="RENOVAÇÃO",
            font=("Arial", 13, "bold"),
            fg=self.cor_zimbra,
            bg="white"
        ).pack(side="left", padx=(12, 12), pady=10)
        self.label_resumo_renovacao = tk.Label(
            cabecalho_renovacao,
            text="0 contrato(s) aguardando tratativa",
            font=("Arial", 11),
            fg="#555555",
            bg="white"
        )
        self.label_resumo_renovacao.pack(side="left", pady=10)
        tk.Button(
            cabecalho_renovacao,
            text="Copiar contratos de renovação",
            command=self.copiar_contratos_renovacao,
            font=("Arial", 9, "bold"),
            bg=self.cor_zimbra,
            fg="white",
            padx=8,
            pady=6
        ).pack(side="right", padx=12, pady=6)

        area_renovacao = tk.Frame(self.aba_renovacao, bg="#F5F6F7")
        area_renovacao.pack(fill="both", expand=True)
        self.tabela_renovacao = ttk.Treeview(area_renovacao, columns=colunas, show="headings")
        for coluna in colunas:
            self.tabela_renovacao.heading(coluna, text=titulos[coluna])
            self.tabela_renovacao.column(coluna, width=larguras[coluna], anchor="center" if coluna in ["contrato", "competencia", "dia"] else "w")
        scroll_renovacao = ttk.Scrollbar(area_renovacao, orient="vertical", command=self.tabela_renovacao.yview)
        self.tabela_renovacao.configure(yscrollcommand=scroll_renovacao.set)
        self.tabela_renovacao.pack(side="left", fill="both", expand=True)
        scroll_renovacao.pack(side="right", fill="y")

        cabecalho_finalizados = tk.Frame(self.aba_finalizados, bg="white", highlightbackground="#D7DCE1", highlightthickness=1)
        cabecalho_finalizados.pack(fill="x", pady=(8, 8))
        tk.Label(
            cabecalho_finalizados,
            text="CONTRATOS FINALIZADOS",
            font=("Arial", 13, "bold"),
            fg="#6B2D73",
            bg="white"
        ).pack(side="left", padx=(12, 12), pady=10)
        self.label_resumo_finalizados = tk.Label(
            cabecalho_finalizados,
            text="Nenhum contrato finalizado no histórico.",
            font=("Arial", 11),
            fg="#555555",
            bg="white"
        )
        self.label_resumo_finalizados.pack(side="left", pady=10)

        area_finalizados = tk.Frame(self.aba_finalizados, bg="#F5F6F7")
        area_finalizados.pack(fill="both", expand=True)
        self.tabela_finalizados = ttk.Treeview(area_finalizados, columns=colunas, show="headings", selectmode="extended")
        for coluna in colunas:
            self.tabela_finalizados.heading(coluna, text=titulos[coluna])
            self.tabela_finalizados.column(coluna, width=larguras[coluna], anchor="center" if coluna in ["contrato", "competencia", "dia"] else "w")
        scroll_finalizados = ttk.Scrollbar(area_finalizados, orient="vertical", command=self.tabela_finalizados.yview)
        self.tabela_finalizados.configure(yscrollcommand=scroll_finalizados.set)
        self.tabela_finalizados.pack(side="left", fill="both", expand=True)
        scroll_finalizados.pack(side="right", fill="y")

        email = tk.Frame(self, bg="white", highlightbackground="#CCCCCC", highlightthickness=1)
        email.pack(fill="x", padx=18, pady=(0, 10))
        tk.Label(
            email, text="E-mail final: copie o texto e preencha destinatário/anexo diretamente no Zimbra.",
            font=("Arial", 11, "bold"), bg="white", fg=self.cor_azul
        ).pack(side="left", padx=12, pady=14)
        self.botao(email, "Copiar texto final", self.copiar_email_final, self.cor_azul)
        self.botao(email, "Copiar texto e abrir Zimbra", self.abrir_zimbra, self.cor_zimbra)
        self.botao(email, "Alterar link Zimbra", self.redefinir_link_zimbra, "#6C757D")

        tk.Label(self, textvariable=self.status_var, font=("Arial", 11, "bold"), bg="#F5F6F7", fg=self.cor_azul).pack(
            anchor="w", padx=18, pady=(0, 12)
        )

    def botao(self, pai, texto, comando, cor):
        tk.Button(pai, text=texto, command=comando, font=("Arial", 9, "bold"), bg=cor, fg="white", padx=8, pady=7).pack(side="left", padx=(0, 7), pady=6)

    def criar_card(self, pai, titulo):
        card = tk.Frame(pai, bg="white", highlightbackground="#DDDDDD", highlightthickness=1)
        card.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Label(card, text=titulo, font=("Arial", 10), fg="#555555", bg="white").pack(anchor="w", padx=12, pady=(8, 1))
        valor = tk.Label(card, text="0", font=("Arial", 18, "bold"), fg=self.cor_azul, bg="white")
        valor.pack(anchor="w", padx=12, pady=(0, 8))
        return valor

    def selecionar_pdf(self):
        caminhos = filedialog.askopenfilenames(
            title="Selecione um ou mais PDFs de anomalias",
            filetypes=[("Arquivo PDF", "*.pdf")]
        )
        if not caminhos:
            return

        total_lidos = 0
        total_novos = 0
        existentes = {chave_anomalia_registro(r) for r in self.registros}

        for caminho in caminhos:
            try:
                registros_pdf = extrair_anomalias_pdf(caminho)
            except Exception as erro:
                messagebox.showerror("Erro", f"Não foi possível ler o PDF:\n{caminho}\n\n{erro}")
                continue

            total_lidos += len(registros_pdf)
            for registro in registros_pdf:
                registro["arquivo_origem"] = os.path.basename(caminho)
                registro["importado_em"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                chave = chave_anomalia_registro(registro)
                if chave not in existentes:
                    self.registros.append(registro)
                    existentes.add(chave)
                    total_novos += 1

        if total_lidos == 0:
            messagebox.showwarning("Sem registros", "Não foram encontradas ocorrências no formato esperado.")
            return

        salvar_historico_anomalias(self.registros)
        self.caminho_pdf = "; ".join(caminhos)
        # Mantém os contratos já finalizados no histórico, mesmo ao adicionar novos PDFs.
        salvar_finalizados_anomalias(self.registros_finalizados)
        self.label_pdf.config(text=f"Histórico local: {len(self.registros)} ocorrência(s) importada(s).")
        self.atualizar_opcoes_competencia()
        self.atualizar_cards()
        self.atualizar_tabela()
        self.status_var.set(
            f"PDF(s) importado(s): {total_lidos} ocorrência(s) lida(s), "
            f"{total_novos} nova(s) adicionada(s) ao histórico."
        )

    def limpar_historico(self):
        if not self.registros:
            messagebox.showinfo("Histórico vazio", "Não há dados de anomalias salvos no histórico.")
            return

        confirmar = messagebox.askyesno(
            "Limpar histórico",
            "Deseja apagar todo o histórico de anomalias salvo neste computador?\n\n"
            "Essa ação não apaga seus PDFs originais."
        )
        if not confirmar:
            return

        self.registros = []
        self.registros_finalizados = []
        salvar_historico_anomalias(self.registros)
        salvar_finalizados_anomalias(self.registros_finalizados)
        self.label_pdf.config(text="Nenhum PDF no histórico.")
        self.atualizar_opcoes_competencia()
        self.atualizar_cards()
        self.atualizar_tabela()
        self.status_var.set("Histórico de anomalias apagado.")

    def aba_renovacao_ativa(self):
        return self.abas_anomalias.index(self.abas_anomalias.select()) == 1

    def aba_finalizados_ativa(self):
        return self.abas_anomalias.index(self.abas_anomalias.select()) == 2

    def chave_registro(self, registro):
        return chave_anomalia_registro(registro)

    def atualizar_opcoes_competencia(self):
        competencias = competencias_anomalias(self.registros + self.registros_finalizados)
        valor_atual = self.competencia_var.get()
        self.combo_competencia_anomalias.config(values=competencias)
        self.competencia_var.set(valor_atual if valor_atual in competencias else "Todas")

    def aplicar_filtro_competencia(self, registros):
        competencia = self.competencia_var.get()
        if competencia == "Todas":
            return registros
        return [r for r in registros if r.get("competencia") == competencia]

    def registros_pendentes(self):
        finalizados = {self.chave_registro(r) for r in self.registros_finalizados}
        pendentes = [r for r in self.registros if self.chave_registro(r) not in finalizados]
        return self.aplicar_filtro_competencia(pendentes)

    def alterar_filtro_por_aba(self, _event=None):
        """Troca automaticamente as opções de categoria conforme a aba aberta."""
        if self.aba_renovacao_ativa():
            self.combo_filtro_anomalias.config(values=self.CATEGORIAS_RENOVACAO, state="readonly")
        elif self.aba_finalizados_ativa():
            self.combo_filtro_anomalias.config(values=["Todos"], state="disabled")
        else:
            self.combo_filtro_anomalias.config(values=self.CATEGORIAS_GERAIS, state="readonly")

        self.filtro_var.set("Todos")
        self.atualizar_tabela()

    def registros_filtrados(self):
        filtro = self.filtro_var.get()
        pendentes = self.registros_pendentes()
        return pendentes if filtro == "Todos" else [r for r in pendentes if r["categoria"] == filtro]


    def registros_renovacao(self):
        categorias_renovacao = {"Aluguel não reajustado", "Seguro fiança vencido"}
        registros = [r for r in self.registros_pendentes() if r["categoria"] in categorias_renovacao]
        filtro = self.filtro_var.get() if self.aba_renovacao_ativa() else "Todos"
        return registros if filtro == "Todos" else [r for r in registros if r["categoria"] == filtro]

    def registros_renovacao_completos(self):
        categorias_renovacao = {"Aluguel não reajustado", "Seguro fiança vencido"}
        return [r for r in self.registros_pendentes() if r["categoria"] in categorias_renovacao]

    def atualizar_cards(self):
        pendentes = self.registros_pendentes()
        self.card_total.config(text=str(len(pendentes)))
        self.card_aluguel.config(text=str(len([r for r in pendentes if r["categoria"] == "Aluguel não reajustado"])))
        self.card_seguro.config(text=str(len([r for r in pendentes if r["categoria"] == "Seguro fiança vencido"])))
        self.card_contratos.config(text=str(len(contratos_anomalia_unicos(pendentes))))

    def atualizar_tabela(self):
        self.atualizar_cards()

        self.tabela.delete(*self.tabela.get_children())
        registros_geral = self.registros_filtrados()
        for item in registros_geral:
            self.tabela.insert("", "end", values=(
                item["categoria"], item["contrato"], item["competencia"],
                item["vencimento"], item["garantia"], item["mensagem"]
            ))

        self.tabela_renovacao.delete(*self.tabela_renovacao.get_children())
        registros_renovacao = self.registros_renovacao()
        for item in registros_renovacao:
            self.tabela_renovacao.insert("", "end", values=(
                item["categoria"], item["contrato"], item["competencia"],
                item["vencimento"], item["garantia"], item["mensagem"]
            ))

        qtd_contratos = len(contratos_anomalia_unicos(registros_renovacao))
        self.label_resumo_renovacao.config(
            text=f"{qtd_contratos} contrato(s) aguardando tratativa | "
                 f"{len(registros_renovacao)} ocorrência(s)"
        )

        finalizados_filtrados = self.aplicar_filtro_competencia(self.registros_finalizados)
        self.tabela_finalizados.delete(*self.tabela_finalizados.get_children())
        for item in finalizados_filtrados:
            self.tabela_finalizados.insert("", "end", values=(
                item["categoria"], item["contrato"], item["competencia"],
                item["vencimento"], item["garantia"], item["mensagem"]
            ))

        contratos_finalizados = len(contratos_anomalia_unicos(finalizados_filtrados))
        self.label_resumo_finalizados.config(
            text=f"{contratos_finalizados} contrato(s) finalizado(s) | "
                 f"{len(finalizados_filtrados)} ocorrência(s)"
            if finalizados_filtrados else "Nenhum contrato finalizado."
        )


    def copiar_texto(self, texto, status):
        self.clipboard_clear()
        self.clipboard_append(texto)
        self.update()
        self.status_var.set(status)
        messagebox.showinfo("Copiado", status)

    def tabela_ativa(self):
        if self.aba_renovacao_ativa():
            return self.tabela_renovacao
        if self.aba_finalizados_ativa():
            return self.tabela_finalizados
        return self.tabela

    def copiar_contrato_selecionado(self):
        tabela_ativa = self.tabela_ativa()
        selecao = tabela_ativa.selection()
        if not selecao:
            messagebox.showwarning("Selecione uma linha", "Selecione um contrato na aba aberta.")
            return
        contrato = tabela_ativa.item(selecao[0], "values")[1]
        self.copiar_texto(contrato, f"Contrato {contrato} copiado.")

    def finalizar_selecionados(self):
        if self.aba_finalizados_ativa():
            messagebox.showinfo("Contratos finalizados", "Os registros desta aba já estão finalizados.")
            return

        tabela = self.tabela_ativa()
        itens_selecionados = tabela.selection()
        if not itens_selecionados:
            messagebox.showwarning(
                "Selecione contratos",
                "Selecione uma ou mais linhas para finalizar."
            )
            return

        chaves_selecionadas = set()
        for item_id in itens_selecionados:
            valores = tabela.item(item_id, "values")
            if valores:
                chaves_selecionadas.add(tuple(valores))

        registros_visiveis = self.registros_renovacao() if self.aba_renovacao_ativa() else self.registros_filtrados()
        ja_finalizados = {self.chave_registro(r) for r in self.registros_finalizados}
        novos_finalizados = [
            r for r in registros_visiveis
            if (
                r["categoria"], r["contrato"], r["competencia"],
                r["vencimento"], r["garantia"], r["mensagem"]
            ) in chaves_selecionadas and self.chave_registro(r) not in ja_finalizados
        ]

        if not novos_finalizados:
            messagebox.showwarning("Sem registros", "Nenhum registro novo foi selecionado para finalizar.")
            return

        data_finalizacao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        for registro in novos_finalizados:
            registro["finalizado_em"] = data_finalizacao

        self.registros_finalizados.extend(novos_finalizados)
        salvar_finalizados_anomalias(self.registros_finalizados)
        self.atualizar_cards()
        self.atualizar_tabela()
        self.abas_anomalias.select(self.aba_finalizados)
        self.status_var.set(f"{len(novos_finalizados)} ocorrência(s) movida(s) para CONTRATOS FINALIZADOS.")
        messagebox.showinfo(
            "Finalizado",
            f"{len(novos_finalizados)} ocorrência(s) foram movida(s) para a aba CONTRATOS FINALIZADOS."
        )

    def copiar_contratos_renovacao(self):
        contratos = contratos_anomalia_unicos(self.registros_renovacao())
        if not contratos:
            messagebox.showwarning("Sem contratos", "Não há contratos para o setor de Renovação.")
            return
        self.copiar_texto(
            "\n".join(contratos),
            f"{len(contratos)} contrato(s) de Renovação copiado(s)."
        )

    def copiar_contratos_filtro(self):
        contratos = contratos_anomalia_unicos(self.registros_filtrados())
        if not contratos:
            messagebox.showwarning("Sem contratos", "Não há contratos no filtro selecionado.")
            return
        self.copiar_texto("\n".join(contratos), f"{len(contratos)} contrato(s) copiado(s).")

    def copiar_mensagem_aluguel(self):
        contratos = contratos_anomalia_unicos([r for r in self.registros if r["categoria"] == "Aluguel não reajustado"])
        if not contratos:
            messagebox.showwarning("Sem registros", "Não existem aluguéis não reajustados no PDF importado.")
            return
        texto = MSG_ANOMALIA_ALUGUEL.format(contratos="\n".join(contratos))
        self.copiar_texto(texto, f"Mensagem copiada com {len(contratos)} contrato(s) não reajustado(s).")

    def copiar_mensagem_seguro(self):
        contratos = contratos_anomalia_unicos([r for r in self.registros if r["categoria"] == "Seguro fiança vencido"])
        if not contratos:
            messagebox.showwarning("Sem registros", "Não existem registros de seguro fiança vencido no PDF importado.")
            return
        texto = MSG_ANOMALIA_SEGURO_FIANCA.format(contratos="\n".join(contratos))
        self.copiar_texto(texto, f"Mensagem de seguro fiança copiada com {len(contratos)} contrato(s).")

    def copiar_email_final(self):
        self.copiar_texto(MSG_ANOMALIA_EMAIL_FINAL, "Texto final do e-mail copiado.")

    def caminho_config_zimbra(self):
        pasta = pasta_config_app()
        return os.path.join(pasta, "config_zimbra.txt")

    def abrir_zimbra(self):
        caminho_config = self.caminho_config_zimbra()
        url = ""
        if os.path.exists(caminho_config):
            with open(caminho_config, "r", encoding="utf-8") as arquivo:
                url = arquivo.read().strip()
        if not url:
            url = simpledialog.askstring(
                "Link do Zimbra",
                "Cole o link de acesso ao Zimbra.\nEsse endereço será salvo para os próximos usos:",
                parent=self
            )
            if not url:
                return
            url = url.strip()
            if not url.startswith(("http://", "https://")):
                url = "https://" + url
            with open(caminho_config, "w", encoding="utf-8") as arquivo:
                arquivo.write(url)
        self.clipboard_clear()
        self.clipboard_append(MSG_ANOMALIA_EMAIL_FINAL)
        self.update()
        webbrowser.open(url, new=2)
        self.status_var.set("Texto final copiado e Zimbra aberto. Inclua destinatário e anexo manualmente.")

    def redefinir_link_zimbra(self):
        caminho_config = self.caminho_config_zimbra()
        if os.path.exists(caminho_config):
            os.remove(caminho_config)
        self.status_var.set("Link do Zimbra removido. Ele será solicitado novamente no próximo acesso.")
        messagebox.showinfo("Link removido", "O link salvo do Zimbra foi removido.")


# ==========================
# TERMO + LOGIN LOCAL + ÁREA ADM + GOOGLE SHEETS ONLINE
# ==========================
import json
import csv
import urllib.request
import urllib.error
import socket

CONFIG_DIR = pasta_config_app()
CONFIG_FILE = os.path.join(CONFIG_DIR, "config_usuario.json")
ACESSOS_FILE = os.path.join(CONFIG_DIR, "acessos_usuarios.json")
TERMO_VERSAO = "2026-05-13"
SENHA_ADMIN = "Reicy"

# IMPORTANTE:
# Depois de criar o Google Apps Script, cole a URL do Web App abaixo.
# Exemplo: GOOGLE_SHEETS_WEB_APP_URL = "https://script.google.com/macros/s/SEU_ID/exec"
GOOGLE_SHEETS_WEB_APP_URL = "https://script.google.com/macros/s/AKfycbyTWj1REo6CXDQYR3M-2AGJMU2qn8bIDHp-Xo3WoewL9XPTSeCRmwS1r4jFBpIaQ8vTzA/exec"


def google_sheets_configurado():
    return (
        GOOGLE_SHEETS_WEB_APP_URL
        and GOOGLE_SHEETS_WEB_APP_URL.startswith("https://script.google.com/macros/s/")
    )


def carregar_config_usuario():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as arquivo:
                return json.load(arquivo)
    except Exception:
        pass
    return {}


def salvar_config_usuario(dados):
    os.makedirs(CONFIG_DIR, exist_ok=True)

    config_atual = carregar_config_usuario()
    config_atual.update(dados)
    config_atual["ultima_atualizacao"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    with open(CONFIG_FILE, "w", encoding="utf-8") as arquivo:
        json.dump(config_atual, arquivo, ensure_ascii=False, indent=4)


def carregar_acessos():
    try:
        if os.path.exists(ACESSOS_FILE):
            with open(ACESSOS_FILE, "r", encoding="utf-8") as arquivo:
                dados = json.load(arquivo)
                if isinstance(dados, list):
                    return dados
    except Exception:
        pass
    return []


def salvar_acessos(acessos):
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(ACESSOS_FILE, "w", encoding="utf-8") as arquivo:
        json.dump(acessos, arquivo, ensure_ascii=False, indent=4)


def montar_dados_acesso(evento, nome="", email=""):
    return {
        "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "evento": evento,
        "nome": nome,
        "email": email,
        "computador": os.environ.get("COMPUTERNAME", socket.gethostname()),
        "usuario_windows": os.environ.get("USERNAME", os.environ.get("USER", "")),
        "versao_termo": TERMO_VERSAO,
        "origem": "Aplicativo Filtro PDF"
    }


def enviar_acesso_online(dados):
    """
    Envia o acesso para o Google Sheets via Google Apps Script.
    Se estiver sem internet ou sem URL configurada, o app continua funcionando normalmente.
    """
    if not google_sheets_configurado():
        return False, "Google Sheets não configurado."

    try:
        payload = json.dumps(dados, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            GOOGLE_SHEETS_WEB_APP_URL,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=10) as resposta:
            retorno = resposta.read().decode("utf-8", errors="ignore")
            return True, retorno
    except Exception as e:
        return False, str(e)


def carregar_acessos_online():
    """
    Lê os acessos direto do Google Sheets via Google Apps Script.
    """
    if not google_sheets_configurado():
        return []

    try:
        with urllib.request.urlopen(GOOGLE_SHEETS_WEB_APP_URL, timeout=10) as resposta:
            retorno = resposta.read().decode("utf-8", errors="ignore")
            dados = json.loads(retorno)
            if isinstance(dados, list):
                return dados
            if isinstance(dados, dict) and isinstance(dados.get("dados"), list):
                return dados.get("dados")
    except Exception:
        pass

    return []


def registrar_acesso(evento, nome="", email=""):
    dados = montar_dados_acesso(evento, nome, email)

    # Salva localmente como backup.
    acessos = carregar_acessos()
    acessos.append(dados)
    salvar_acessos(acessos)

    # Também tenta enviar para o Google Sheets online.
    enviar_acesso_online(dados)


def exportar_acessos_csv(acessos=None):
    if acessos is None:
        acessos = carregar_acessos_online() or carregar_acessos()

    if not acessos:
        messagebox.showwarning("Sem registros", "Ainda não existe nenhum acesso registrado.")
        return

    caminho = filedialog.asksaveasfilename(
        title="Salvar relatório de acessos",
        defaultextension=".csv",
        filetypes=[("Arquivo CSV", "*.csv")]
    )

    if not caminho:
        return

    colunas = [
        "data_hora",
        "evento",
        "nome",
        "email",
        "computador",
        "usuario_windows",
        "versao_termo",
        "origem"
    ]

    try:
        with open(caminho, "w", encoding="utf-8-sig", newline="") as arquivo:
            writer = csv.DictWriter(arquivo, fieldnames=colunas, delimiter=";")
            writer.writeheader()
            for item in acessos:
                writer.writerow({col: item.get(col, "") for col in colunas})

        messagebox.showinfo("Exportado", f"Relatório salvo com sucesso em:\n{caminho}")
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível exportar o relatório.\n\n{e}")


def exibir_area_admin(root):
    janela_senha = tk.Toplevel(root)
    janela_senha.title("Área ADM")
    janela_senha.geometry("380x190")
    janela_senha.configure(bg="#F5F6F7")
    janela_senha.resizable(False, False)

    senha_var = tk.StringVar()
    autorizado = {"valor": False}

    tk.Label(
        janela_senha,
        text="Acesso do administrador",
        font=("Arial", 16, "bold"),
        fg="#1F4E78",
        bg="#F5F6F7"
    ).pack(pady=(18, 10))

    frame = tk.Frame(janela_senha, bg="#F5F6F7")
    frame.pack(fill="x", padx=28)

    tk.Label(frame, text="Senha ADM", font=("Arial", 11, "bold"), bg="#F5F6F7", fg="#1F4E78").pack(anchor="w")
    campo_senha = tk.Entry(frame, textvariable=senha_var, font=("Arial", 12), show="*")
    campo_senha.pack(fill="x", pady=(2, 12), ipady=5)
    campo_senha.focus_set()

    def validar_senha():
        if senha_var.get().strip() == SENHA_ADMIN:
            autorizado["valor"] = True
            janela_senha.destroy()
        else:
            messagebox.showerror("Senha incorreta", "A senha de administrador está incorreta.")

    tk.Button(
        janela_senha,
        text="Entrar",
        command=validar_senha,
        font=("Arial", 11, "bold"),
        bg="#2F9254",
        fg="white",
        width=14
    ).pack(pady=(2, 12))

    janela_senha.bind("<Return>", lambda event: validar_senha())
    janela_senha.grab_set()
    root.wait_window(janela_senha)

    if not autorizado["valor"]:
        return

    acessos_online = carregar_acessos_online()
    acessos_locais = carregar_acessos()

    # Prioriza o online. Se não tiver online configurado ou falhar, mostra o local.
    if acessos_online:
        acessos = acessos_online
        origem_titulo = "Acessos online - Google Sheets"
    else:
        acessos = acessos_locais
        origem_titulo = "Acessos locais deste computador"

    janela = tk.Toplevel(root)
    janela.title("Área ADM - Acessos do Aplicativo")
    janela.geometry("1060x560")
    janela.configure(bg="#F5F6F7")

    tk.Label(
        janela,
        text=origem_titulo,
        font=("Arial", 18, "bold"),
        fg="#1F4E78",
        bg="#F5F6F7"
    ).pack(pady=(15, 4))

    status_google = "Google Sheets configurado" if google_sheets_configurado() else "Google Sheets ainda não configurado no código"
    tk.Label(
        janela,
        text=f"Total de registros exibidos: {len(acessos)} | {status_google}",
        font=("Arial", 11),
        fg="#555555",
        bg="#F5F6F7"
    ).pack(pady=(0, 10))

    frame_texto = tk.Frame(janela, bg="#F5F6F7")
    frame_texto.pack(fill="both", expand=True, padx=18, pady=(0, 10))

    texto = tk.Text(frame_texto, font=("Consolas", 10), bg="white", fg="black", wrap="none")
    texto.pack(side="left", fill="both", expand=True)

    scroll_y = tk.Scrollbar(frame_texto, orient="vertical", command=texto.yview)
    scroll_y.pack(side="right", fill="y")
    texto.configure(yscrollcommand=scroll_y.set)

    if not acessos:
        if google_sheets_configurado():
            texto.insert("end", "Nenhum acesso encontrado. Verifique se o Google Apps Script foi publicado corretamente.\n")
        else:
            texto.insert("end", "Nenhum acesso registrado neste computador.\n")
    else:
        cabecalho = f"{'DATA/HORA':<20} | {'EVENTO':<28} | {'NOME':<28} | {'E-MAIL':<32} | {'PC':<18} | {'USUÁRIO WINDOWS':<18} | {'TERMO':<10}\n"
        texto.insert("end", cabecalho)
        texto.insert("end", "-" * 170 + "\n")

        for item in acessos:
            linha = (
                f"{item.get('data_hora', ''):<20} | "
                f"{item.get('evento', ''):<28} | "
                f"{item.get('nome', ''):<28} | "
                f"{item.get('email', ''):<32} | "
                f"{item.get('computador', ''):<18} | "
                f"{item.get('usuario_windows', ''):<18} | "
                f"{item.get('versao_termo', ''):<10}\n"
            )
            texto.insert("end", linha)

    texto.config(state="disabled")

    botoes = tk.Frame(janela, bg="#F5F6F7")
    botoes.pack(pady=(0, 15))

    def atualizar_lista():
        janela.destroy()
        exibir_area_admin(root)

    tk.Button(
        botoes,
        text="Atualizar",
        command=atualizar_lista,
        font=("Arial", 11, "bold"),
        bg="#2F9254",
        fg="white",
        width=16
    ).pack(side="left", padx=8)

    tk.Button(
        botoes,
        text="Exportar CSV",
        command=lambda: exportar_acessos_csv(acessos),
        font=("Arial", 11, "bold"),
        bg="#1F4E78",
        fg="white",
        width=16
    ).pack(side="left", padx=8)

    tk.Button(
        botoes,
        text="Fechar",
        command=janela.destroy,
        font=("Arial", 11, "bold"),
        bg="#6C757D",
        fg="white",
        width=16
    ).pack(side="left", padx=8)


def exibir_termo_consentimento(root, config_salva=None):
    if config_salva is None:
        config_salva = {}

    termo = """
TERMO DE CONSENTIMENTO E USO DO APLICATIVO

Este aplicativo foi desenvolvido e pertence a:

Emanuel Monteiro Silva

Ao utilizar este sistema, o usuário declara estar ciente de que:

1. O aplicativo é de propriedade de Emanuel Monteiro Silva.

2. É proibida a cópia, venda, distribuição, modificação, compartilhamento ou reprodução deste aplicativo, total ou parcial, sem autorização prévia do proprietário.

3. O aplicativo foi desenvolvido para auxiliar no processamento de arquivos PDF e geração de planilhas em Excel, conforme sua finalidade original.

4. O proprietário não se responsabiliza por alterações feitas no código, uso indevido, arquivos incorretos inseridos no sistema ou interpretações equivocadas dos dados gerados.

5. O uso deste aplicativo implica a aceitação integral deste termo.

Ao clicar em “Aceito”, o usuário confirma que leu, compreendeu e concorda com as condições acima.
"""

    janela = tk.Toplevel(root)
    janela.title("Termo de Consentimento")
    janela.geometry("760x720")
    janela.configure(bg="#F5F6F7")
    janela.resizable(False, False)

    resultado = {"autorizado": False}

    nome_var = tk.StringVar(value=config_salva.get("nome", ""))
    email_var = tk.StringVar(value=config_salva.get("email", ""))
    senha_var = tk.StringVar(value=config_salva.get("senha", ""))

    tk.Label(
        janela,
        text="Termo de Consentimento e Uso",
        font=("Arial", 18, "bold"),
        fg="#1F4E78",
        bg="#F5F6F7"
    ).pack(pady=(15, 8))

    texto = tk.Text(
        janela,
        font=("Arial", 11),
        wrap="word",
        height=17,
        padx=12,
        pady=12
    )
    texto.pack(fill="both", expand=True, padx=20, pady=(5, 10))
    texto.insert("1.0", termo)
    texto.config(state="disabled")

    frame_campos = tk.Frame(janela, bg="#F5F6F7")
    frame_campos.pack(fill="x", padx=22, pady=(5, 5))

    tk.Label(frame_campos, text="Nome completo", font=("Arial", 11, "bold"), bg="#F5F6F7", fg="#1F4E78").pack(anchor="w")
    tk.Entry(frame_campos, textvariable=nome_var, font=("Arial", 12)).pack(fill="x", pady=(2, 8), ipady=5)

    tk.Label(frame_campos, text="E-mail", font=("Arial", 11, "bold"), bg="#F5F6F7", fg="#1F4E78").pack(anchor="w")
    tk.Entry(frame_campos, textvariable=email_var, font=("Arial", 12)).pack(fill="x", pady=(2, 8), ipady=5)

    tk.Label(frame_campos, text="Senha", font=("Arial", 11, "bold"), bg="#F5F6F7", fg="#1F4E78").pack(anchor="w")
    tk.Entry(frame_campos, textvariable=senha_var, font=("Arial", 12), show="*").pack(fill="x", pady=(2, 8), ipady=5)

    frame_botoes = tk.Frame(janela, bg="#F5F6F7")
    frame_botoes.pack(pady=15)

    def aceitar():
        nome = nome_var.get().strip()
        email = email_var.get().strip()
        senha = senha_var.get().strip()

        if not nome:
            messagebox.showerror("Campo obrigatório", "Informe o nome completo.")
            return

        if not email:
            messagebox.showerror("Campo obrigatório", "Informe o e-mail.")
            return

        if not senha:
            messagebox.showerror("Campo obrigatório", "Informe a senha.")
            return

        agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        salvar_config_usuario({
            "termo_aceito": True,
            "termo_versao": TERMO_VERSAO,
            "nome": nome,
            "email": email,
            "senha": senha,
            "primeiro_aceite_em": config_salva.get("primeiro_aceite_em") or agora,
            "ultimo_acesso_em": agora
        })

        registrar_acesso("Aceitou termo e acessou", nome, email)

        resultado["autorizado"] = True
        janela.destroy()

    def recusar():
        resultado["autorizado"] = False
        janela.destroy()

    tk.Button(
        frame_botoes,
        text="Aceito",
        command=aceitar,
        font=("Arial", 12, "bold"),
        bg="#2F9254",
        fg="white",
        width=15
    ).pack(side="left", padx=10)

    tk.Button(
        frame_botoes,
        text="Não aceito",
        command=recusar,
        font=("Arial", 12, "bold"),
        bg="#A94442",
        fg="white",
        width=15
    ).pack(side="left", padx=10)

    tk.Button(
        frame_botoes,
        text="Área ADM",
        command=lambda: exibir_area_admin(root),
        font=("Arial", 12, "bold"),
        bg="#343A40",
        fg="white",
        width=15
    ).pack(side="left", padx=10)

    janela.protocol("WM_DELETE_WINDOW", recusar)
    janela.grab_set()
    root.wait_window(janela)

    return resultado["autorizado"]


def exibir_login_usuario(root, config_salva=None):
    if config_salva is None:
        config_salva = {}

    janela = tk.Toplevel(root)
    janela.title("Acesso ao aplicativo")
    janela.geometry("500x360")
    janela.configure(bg="#F5F6F7")
    janela.resizable(False, False)

    resultado = {"autorizado": False}

    email_var = tk.StringVar(value=config_salva.get("email", ""))
    senha_var = tk.StringVar(value=config_salva.get("senha", ""))

    tk.Label(
        janela,
        text="Acesso ao aplicativo",
        font=("Arial", 18, "bold"),
        fg="#1F4E78",
        bg="#F5F6F7"
    ).pack(pady=(22, 6))

    tk.Label(
        janela,
        text="O termo de uso já foi aceito neste computador.",
        font=("Arial", 10),
        fg="#555555",
        bg="#F5F6F7"
    ).pack(pady=(0, 18))

    frame = tk.Frame(janela, bg="#F5F6F7")
    frame.pack(fill="x", padx=30)

    tk.Label(frame, text="E-mail", font=("Arial", 11, "bold"), bg="#F5F6F7", fg="#1F4E78").pack(anchor="w")
    tk.Entry(frame, textvariable=email_var, font=("Arial", 12)).pack(fill="x", pady=(2, 10), ipady=5)

    tk.Label(frame, text="Senha", font=("Arial", 11, "bold"), bg="#F5F6F7", fg="#1F4E78").pack(anchor="w")
    tk.Entry(frame, textvariable=senha_var, font=("Arial", 12), show="*").pack(fill="x", pady=(2, 12), ipady=5)

    def entrar():
        email = email_var.get().strip()
        senha = senha_var.get().strip()

        if not email:
            messagebox.showerror("Campo obrigatório", "Informe o e-mail.")
            return

        if not senha:
            messagebox.showerror("Campo obrigatório", "Informe a senha.")
            return

        agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        nome = config_salva.get("nome", "")

        salvar_config_usuario({
            "email": email,
            "senha": senha,
            "ultimo_acesso_em": agora
        })

        registrar_acesso("Acessou o aplicativo", nome, email)

        resultado["autorizado"] = True
        janela.destroy()

    def cancelar():
        resultado["autorizado"] = False
        janela.destroy()

    frame_botoes = tk.Frame(janela, bg="#F5F6F7")
    frame_botoes.pack(pady=12)

    tk.Button(
        frame_botoes,
        text="Entrar",
        command=entrar,
        font=("Arial", 12, "bold"),
        bg="#2F9254",
        fg="white",
        width=13
    ).pack(side="left", padx=8)

    tk.Button(
        frame_botoes,
        text="Cancelar",
        command=cancelar,
        font=("Arial", 12, "bold"),
        bg="#6C757D",
        fg="white",
        width=13
    ).pack(side="left", padx=8)

    tk.Button(
        frame_botoes,
        text="Área ADM",
        command=lambda: exibir_area_admin(root),
        font=("Arial", 12, "bold"),
        bg="#343A40",
        fg="white",
        width=13
    ).pack(side="left", padx=8)

    janela.protocol("WM_DELETE_WINDOW", cancelar)
    janela.grab_set()
    root.wait_window(janela)

    return resultado["autorizado"]


# ==========================
# MENU INICIAL COM ÍCONES DE MÓDULOS
# ==========================

class MenuModulos:
    def __init__(self, root):
        self.root = root
        self.root.title("Central de Ferramentas | Emanuel | V6 Finalizados")
        self.root.geometry("1040x660")
        self.root.minsize(900, 560)
        self.root.configure(bg="#F5F6F7")

        self.cor_azul = "#1F4E78"
        self.cor_verde = "#2F9254"
        self.cor_anomalias = "#087E8B"
        self.cor_fundo = "#F5F6F7"
        self.cor_caixa = "#FFFFFF"

        self.montar_tela()

    def montar_tela(self):
        topo = tk.Frame(self.root, bg=self.cor_azul, height=110)
        topo.pack(fill="x")
        topo.pack_propagate(False)

        tk.Label(
            topo,
            text="Central de Ferramentas",
            font=("Arial", 27, "bold"),
            fg="white",
            bg=self.cor_azul
        ).pack(anchor="w", padx=42, pady=(22, 0))

        tk.Label(
            topo,
            text="Selecione o módulo que deseja utilizar",
            font=("Arial", 12),
            fg="white",
            bg=self.cor_azul
        ).pack(anchor="w", padx=44, pady=(2, 0))

        corpo = tk.Frame(self.root, bg=self.cor_fundo)
        corpo.pack(fill="both", expand=True, padx=42, pady=34)

        tk.Label(
            corpo,
            text="Módulos disponíveis",
            font=("Arial", 17, "bold"),
            fg=self.cor_azul,
            bg=self.cor_fundo
        ).pack(anchor="w", pady=(0, 22))

        cards = tk.Frame(corpo, bg=self.cor_fundo)
        cards.pack(fill="x")

        self.criar_card_modulo(
            cards,
            icone="📄",
            titulo="Filtro de PDF para Excel",
            descricao="Processar repasses, contratos AJ,\nrescisões e prioridade de cobrança.",
            cor_botao=self.cor_azul,
            comando=self.abrir_filtro_pdf
        ).pack(side="left", fill="both", expand=True, padx=(0, 16))

        self.criar_card_modulo(
            cards,
            icone="⚠",
            titulo="Anomalias e Reajustes",
            descricao="Conferir reajustes, seguros,\ncopiar contratos e abrir o Zimbra.",
            cor_botao=self.cor_anomalias,
            comando=self.abrir_anomalias
        ).pack(side="left", fill="both", expand=True, padx=(0, 16))

        self.criar_card_modulo(
            cards,
            icone="🔒",
            titulo="Área ADM",
            descricao="Consultar registros de acesso\ne exportar relatórios.",
            cor_botao="#343A40",
            comando=self.abrir_area_admin
        ).pack(side="left", fill="both", expand=True)

        rodape = tk.Frame(self.root, bg="#E9ECEF", height=55)
        rodape.pack(fill="x", side="bottom")
        rodape.pack_propagate(False)
        tk.Label(
            rodape,
            text="Selecione um ícone para abrir a ferramenta desejada.",
            font=("Arial", 11),
            fg="#555555",
            bg="#E9ECEF"
        ).pack(anchor="w", padx=42, pady=18)

    def criar_card_modulo(self, pai, icone, titulo, descricao, cor_botao, comando):
        card = tk.Frame(
            pai,
            bg=self.cor_caixa,
            highlightbackground="#D7DCE1",
            highlightthickness=1,
            padx=18,
            pady=18
        )

        tk.Label(
            card,
            text=icone,
            font=("Segoe UI Emoji", 42),
            fg=cor_botao,
            bg=self.cor_caixa
        ).pack(pady=(6, 8))

        tk.Label(
            card,
            text=titulo,
            font=("Arial", 14, "bold"),
            fg=self.cor_azul,
            bg=self.cor_caixa,
            justify="center"
        ).pack(pady=(0, 9))

        tk.Label(
            card,
            text=descricao,
            font=("Arial", 10),
            fg="#555555",
            bg=self.cor_caixa,
            justify="center"
        ).pack(pady=(0, 18))

        tk.Button(
            card,
            text="ABRIR",
            command=comando,
            font=("Arial", 11, "bold"),
            bg=cor_botao,
            fg="white",
            activebackground=cor_botao,
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            width=20,
            pady=10
        ).pack(pady=(0, 6))

        return card

    def aplicar_icone_janela(self, janela):
        try:
            janela.iconbitmap("icone.ico")
        except Exception:
            pass

    def ocultar_menu(self):
        self.root.withdraw()

    def voltar_ao_menu(self, janela):
        try:
            janela.destroy()
        finally:
            self.root.deiconify()
            self.root.lift()
            self.root.focus_force()

    def abrir_area_admin(self):
        self.ocultar_menu()
        try:
            exibir_area_admin(self.root)
        finally:
            self.root.deiconify()
            self.root.lift()

    def abrir_filtro_pdf(self):
        self.ocultar_menu()
        janela = tk.Toplevel(self.root)
        self.aplicar_icone_janela(janela)
        AppFiltroPDF(janela)
        janela.protocol("WM_DELETE_WINDOW", lambda: self.voltar_ao_menu(janela))

    def abrir_anomalias(self):
        self.ocultar_menu()
        janela = JanelaAnomalias(self.root)
        self.aplicar_icone_janela(janela)
        janela.protocol("WM_DELETE_WINDOW", lambda: self.voltar_ao_menu(janela))


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()

    try:
        root.iconbitmap("icone.ico")
    except Exception:
        pass

    config_usuario = carregar_config_usuario()

    if config_usuario.get("termo_aceito") is True:
        autorizado = exibir_login_usuario(root, config_usuario)
    else:
        autorizado = exibir_termo_consentimento(root, config_usuario)

    if not autorizado:
        messagebox.showwarning(
            "Acesso negado",
            "Para utilizar o aplicativo, é necessário informar os dados de acesso e aceitar o termo de consentimento."
        )
        root.destroy()
    else:
        root.deiconify()
        app = MenuModulos(root)
        root.mainloop()
